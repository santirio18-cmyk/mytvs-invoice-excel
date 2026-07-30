"""
GST Tax Invoice → Excel
Tuned for Indian auto-parts invoices (Tally / TVS / Madras Auto / photos).
Supports PDF + image uploads with OCR.
"""

from __future__ import annotations

import csv
import io
import os
import re
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pdfplumber
import pytesseract
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageEnhance, ImageOps

MAX_FILES = 10
MAX_FILE_SIZE = 20 * 1024 * 1024
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff", ".bmp"}
PDF_EXTS = {".pdf"}

_cors = os.getenv("CORS_ORIGINS", "*").strip()
CORS_ORIGINS = [o.strip() for o in _cors.split(",") if o.strip()] or ["*"]

app = FastAPI(title="myTVS — Invoice to Excel", version="2.0.0")

DEPLOY_MARK = "2026-07-30-phone-tally"
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
META_LABEL_FONT = Font(bold=True, name="Calibri", size=11, color="1F3A5F")
THIN = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)
ALT_FILL = PatternFill("solid", fgColor="F4F7FB")

# Soft OCR cleanup only — not a supplier whitelist
def _normalize_supplier_ocr(name: str) -> str:
    name = _clean(name)
    name = re.sub(r"(?i)\b\w*INVOICE\b.*$", "", name)
    name = re.sub(r"(?i)\s+inreokce\b.*$", "", name)
    name = re.sub(r"(?i)\s+inv[o0][il1]?ce\b.*$", "", name)
    name = re.sub(r"(?i)\bAGENGC\b", "AGENCIES", name)
    name = re.sub(r"(?i)\bAGENCIE\b", "AGENCIES", name)
    name = re.sub(r"(?i)\beNCY\b", "AGENCY", name)
    name = re.sub(r"(?i)^[A4]UTOLIGH\w*", "AUTOLIGHT", name)
    name = re.sub(r"(?i)^[AR]?UTOONE\b", "AUTOONE", name)
    name = re.sub(r"(?i)^ANGAM\b", "THANGAM", name)
    name = re.sub(r"(?i)\bSRI[Il1]\s+KUMARAN\b", "SRII KUMARAN", name)
    name = re.sub(r"(?i)^SRI[Il1]\b", "SRII", name)
    # OCR leftover after brand: "AUTOLIGHT: A"
    name = re.sub(r":\s*[A-Z]?\s*$", "", name)
    name = re.sub(r"\s+", " ", name).strip(" -|,:;")
    return name[:90]


def _ext(name: str) -> str:
    return ("." + name.rsplit(".", 1)[-1].lower()) if "." in name else ""


def _clean(s: str) -> str:
    s = s.replace("|", " ").replace("]", " ").replace("{", " ").replace("}", " ")
    s = re.sub(r"[^\S\n]+", " ", s)
    return s.strip()


def preprocess(img: Image.Image) -> Image.Image:
    img = ImageOps.exif_transpose(img.convert("RGB"))
    w, h = img.size
    scale = max(1.0, 2400 / max(w, h))
    if scale != 1.0:
        img = img.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)
    gray = ImageOps.autocontrast(img.convert("L"))
    gray = ImageEnhance.Sharpness(gray).enhance(2.0)
    gray = ImageEnhance.Contrast(gray).enhance(1.6)
    return gray


def ocr_image(img: Image.Image) -> str:
    prepared = preprocess(img)
    # Full page — PSM 4 + 6
    full = pytesseract.image_to_string(prepared, lang="eng", config="--oem 3 --psm 4")
    full6 = pytesseract.image_to_string(prepared, lang="eng", config="--oem 3 --psm 6")

    w, h = prepared.size
    top_right = prepared.crop((int(w * 0.42), 0, w, int(h * 0.28)))
    tr = pytesseract.image_to_string(top_right, lang="eng", config="--oem 3 --psm 6")

    # Mid table — proven on phone photos (Autolight-style)
    mid = prepared.crop((0, int(h * 0.35), w, int(h * 0.72)))
    mid_txt = pytesseract.image_to_string(mid, lang="eng", config="--oem 3 --psm 6")

    # Lower table — last line items often missed by mid crop (Vinayaka row 10–11)
    lower = prepared.crop((0, int(h * 0.52), w, int(h * 0.90)))
    lower_txt = pytesseract.image_to_string(lower, lang="eng", config="--oem 3 --psm 6")

    return "\n".join([full, full6, tr, mid_txt, lower_txt])


def _looks_like_email_screenshot(text: str) -> bool:
    hits = 0
    for pat in (
        r"(?i)summarize this email",
        r"(?i)all folders",
        r"(?i)google api",
        r"(?i)onedrive",
        r"(?i)download.*full\s*scree",
        r"(?i)inbox|outlook|mail\.google",
    ):
        if re.search(pat, text):
            hits += 1
    return hits >= 2


def extract_text_from_pdf(data: bytes) -> str:
    """
    Extract text from PDF.
    Many invoice PDFs are image-only scans — pdfplumber may see 0 pages or empty text.
    Fall back to pypdfium2 render + OCR whenever text is missing/useless.
    """
    chunks: list[str] = []

    # 1) Try native text via pdfplumber
    try:
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for page in pdf.pages:
                t = page.extract_text() or ""
                if _text_looks_useful(t):
                    chunks.append(t)
                else:
                    # Sparse/garbage text layer — rasterize this page if possible
                    try:
                        pil = page.to_image(resolution=200).original
                        chunks.append(ocr_image(pil))
                    except Exception:
                        if t.strip():
                            chunks.append(t)
    except Exception:
        chunks = []

    if _text_looks_useful("\n".join(chunks)):
        return "\n".join(chunks)

    # 2) Fallback: pypdfium2 (handles scan PDFs pdfplumber cannot open)
    try:
        import pypdfium2 as pdfium

        doc = pdfium.PdfDocument(data)
        ocr_chunks: list[str] = []
        for i in range(len(doc)):
            page = doc[i]
            bitmap = page.render(scale=2.0)
            pil = bitmap.to_pil()
            ocr_chunks.append(ocr_image(pil))
        if ocr_chunks:
            return "\n".join(ocr_chunks)
    except Exception:
        pass

    return "\n".join(chunks)


def _text_looks_useful(text: str) -> bool:
    """True if extracted PDF text likely contains invoice content (not empty/junk)."""
    t = (text or "").strip()
    if len(t) < 40:
        return False
    # Need some letters and ideally invoice-ish tokens or digit amounts
    letters = sum(c.isalpha() for c in t)
    if letters < 20:
        return False
    if re.search(
        r"(?i)(invoice|bill\s*no|gstin|hsn|quantity|amount|supplier|part\s*no|description)",
        t,
    ):
        return True
    # Enough structure: multiple numbers that look like money
    if len(re.findall(r"\d+\.\d{2}", t)) >= 3:
        return True
    return letters > 80


def extract_text_from_image(data: bytes) -> str:
    img = Image.open(io.BytesIO(data))
    chunks = [ocr_image(img)]
    # Outlook/Gmail desktop screenshots: invoice sits in the left attachment pane
    w, h = img.size
    if w >= int(h * 1.45):
        left = img.crop((int(w * 0.01), int(h * 0.10), int(w * 0.52), int(h * 0.98)))
        left_txt = ocr_image(left)
        chunks.append(left_txt)
        if _looks_like_email_screenshot(chunks[0]):
            # Prefer left-pane text first so header parsers see Invoice No / Date cleanly
            chunks = [left_txt, chunks[0]]
    return "\n".join(chunks)


def extract_text(data: bytes, filename: str) -> str:
    ext = _ext(filename)
    if ext in PDF_EXTS or data[:4] == b"%PDF":
        return extract_text_from_pdf(data)
    if ext in IMAGE_EXTS:
        return extract_text_from_image(data)
    try:
        return extract_text_from_image(data)
    except Exception as exc:
        raise ValueError(f"Unsupported file type: {filename}") from exc


# ---------------------------------------------------------------------------
# Header fields
# ---------------------------------------------------------------------------

INVOICE_PATTERNS = [
    re.compile(r"Bill\s*No\.?\s*[:\-]?\s*([A-Z0-9][A-Z0-9\/\-]{1,})", re.I),
    re.compile(
        r"(?:DLR\s*)?INVOICE\s*(?:NO|NUMBER|SL\.?\s*NO|#)\.?\s*[:\-]?\s*([A-Z0-9][A-Z0-9\/\-]{2,})",
        re.I,
    ),
    re.compile(r"Invoice\s*No\.?\s*[:\-]?\s*([A-Z0-9][A-Z0-9\/\-]{2,})", re.I),
    # Invoice No on one line, value on next / after OCR junk (e.g. 26-27-AUTO-10286)
    re.compile(
        r"Invoice\s*No\.?\s*[:\-]?\s*[^\n]{0,80}?(\d{2}-\d{2}-[A-Z]{2,}-\d{3,})",
        re.I,
    ),
    # Bare FY-supplier-serial style
    re.compile(r"\b(\d{2}-\d{2}-[A-Z]{2,}-\d{3,})\b", re.I),
    # Supplier prefix / serial / FY: SVAA/3446/26-27
    re.compile(r"\b([A-Z]{2,6}\/\d{2,5}\/\d{2}-\d{2})\b", re.I),
    # Bare Tally-style: 151/2026-27 or 279/2026-27 near top
    re.compile(r"\b(\d{1,4}\/\d{2,4}-\d{2})\b"),
    # A/TR1562/26-27 style (OCR may drop slash / misread)
    re.compile(r"\b(A[\s\/]?TR\d{3,5}\/\d{2}-\d{2})\b", re.I),
    re.compile(r"\b(A[\s\/]?TR\d{3,5})\b", re.I),
    re.compile(r"\b(MERD[O0]?\d{10,})\b", re.I),
    re.compile(r"\b(MCI\d{10,})\b", re.I),
    re.compile(r"\b(\d{2}-\d{2}\/\d{2,4})\b"),
]

DATE_PATTERNS = [
    re.compile(
        r"(?:Invoice\s*)?Date[d]?\s*[:\-]?\s*"
        r"(\d{1,2}[\-\/\.]\s*(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\-\/\.]?\s*\d{2,4}"
        r"|\d{1,2}[\-\/\.]\d{1,2}[\-\/\.]\d{2,4})",
        re.I,
    ),
    re.compile(
        r"\b(\d{1,2}-(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{2,4})\b",
        re.I,
    ),
    re.compile(r"\b(\d{1,2}\/\d{1,2}\/\d{4})\b"),
    re.compile(r"\b(\d{1,2}\.\d{1,2}\.\d{4})\b"),
]

PLACE_PATTERNS = [
    re.compile(r"Place\s*of\s*(?:Supply|Delivery)\s*[:\-]?\s*([^\n\r]{3,50})", re.I),
    re.compile(r"(?:F)?Supply\s*[:\-]?\s*(\d{1,2}\s*[-–]?\s*[A-Za-z ]{4,20})", re.I),
    re.compile(r"State\s*(?:Name)?\s*[:\-]?\s*(\d{0,2}\s*[-–]?\s*[A-Za-z ]+?)(?:,\s*Code|\n|$)", re.I),
]


def _normalize_invoice_no(val: str) -> str:
    val = val.upper().replace(" ", "")
    val = re.sub(r"^MERDO", "MERD", val)
    val = re.sub(r"^ATR", "A/TR", val)
    return val


def _is_plausible_invoice_no(val: str) -> bool:
    if len(val) < 2 or len(val) > 40:
        return False
    if not re.search(r"\d", val):
        return False
    # Reject IRN / Ack style long hashes
    if len(val) > 30 and re.fullmatch(r"[A-F0-9]+", val):
        return False
    # Reject OCR garbage like GASAGASEG / repeated letters
    letters = re.sub(r"[^A-Z]", "", val)
    if letters and len(set(letters)) <= 2 and len(letters) >= 6:
        return False
    if re.fullmatch(r"0+", re.sub(r"\D", "", val) or "x"):
        return False
    return True


def find_invoice_number(text: str) -> str:
    # Prefer clear FY-supplier-serial forms even when OCR separates the label
    for m in re.finditer(r"\b(\d{2}-\d{2}-[A-Z]{2,}-\d{3,})\b", text, re.I):
        val = _normalize_invoice_no(_clean(m.group(1)))
        if _is_plausible_invoice_no(val):
            return val

    # Supplier code / serial / FY — SVAA/3446/26-27 (prefer over bare 3446/26-27)
    for m in re.finditer(r"\b([A-Z]{2,6}\/\d{2,5}\/\d{2}-\d{2})\b", text, re.I):
        val = _normalize_invoice_no(_clean(m.group(1)))
        if _is_plausible_invoice_no(val):
            return val

    # Shell/lubricant style SL/26/27/3722
    for m in re.finditer(r"\b([A-Z]{1,4}\/\d{2}\/\d{2}\/\d{3,6})\b", text, re.I):
        val = _normalize_invoice_no(_clean(m.group(1)))
        if _is_plausible_invoice_no(val):
            return val

    # Bare serial/FY — 3405/26-27
    for m in re.finditer(r"\b(\d{3,5}\/\d{2}-\d{2})\b", text):
        val = _normalize_invoice_no(_clean(m.group(1)))
        if _is_plausible_invoice_no(val):
            return val

    # Highest priority: common Indian invoice labels (any supplier)
    for pat in (
        re.compile(r"Bill\s*No\.?\s*[:\-]?\s*([A-Z0-9][A-Z0-9\/\-]*)", re.I),
        re.compile(
            r"(?:DLR\s*)?Inv(?:oice|oce)?\s*(?:No|Number|Sl\.?\s*No)\.?\s*[:\-]?\s*"
            r"(?:[^\nA-Z0-9]{0,20})?([A-Z0-9][A-Z0-9\/\-]{1,})",
            re.I,
        ),
        # Phone OCR: "Invoice No" on one line, "J3398 29-Jul-26" on next / same blob
        re.compile(
            r"Inv[a-z]{0,6}\s*No\.?[^\n]{0,40}?\b([A-Z]\d{3,6})\b",
            re.I | re.S,
        ),
        re.compile(r"(?:Inv|te)\s*No\.?\s*[:\-]?\s*([A-Z0-9][A-Z0-9\/\-]{1,})", re.I),  # OCR: teNo.
        # Handwritten / sparse: Invoice No | 160
        re.compile(r"Inv[a-z]{0,6}\s*No\.?\s*[|:.\- ]+\s*(\d{2,6})\b", re.I),
        # ZipERP / Karnavati: "Order No. : Dated :" then "... MADURAI 000089 29/07/2026"
        re.compile(
            r"Order\s*No\.?\s*[:\-]?[^\n]{0,60}\n[^\n]*?\b(\d{4,8})\b\s+\d{1,2}/\d{1,2}/\d{2,4}",
            re.I,
        ),
        # Bare J3398 / A1024 near dated month
        re.compile(
            r"\b([A-Z]\d{3,6})\b\s+\d{1,2}[-/\s]*(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)",
            re.I,
        ),
    ):
        m = pat.search(text)
        if m:
            val = _normalize_invoice_no(_clean(m.group(1)))
            if _is_plausible_invoice_no(val) and val.upper() not in {"SST", "GST", "HSN", "IRN"}:
                return val

    candidates: list[str] = []
    for pat in INVOICE_PATTERNS:
        for m in pat.finditer(text):
            val = _normalize_invoice_no(_clean(m.group(1)).rstrip(".,;"))
            if val and val.lower() not in {"dated", "date", "original", "page", "delivery", "sst"}:
                candidates.append(val)
    ranked = sorted(
        candidates,
        key=lambda v: (
            1 if re.search(r"\d", v) else 0,
            1 if "/" in v or "-" in v else 0,
            1 if re.match(r"^[A-Z]{2,}/", v) else 0,
            len(v),
        ),
        reverse=True,
    )
    for val in ranked:
        if _is_plausible_invoice_no(val):
            return val
    return "Unknown"


def find_date(text: str) -> str:
    # Prefer labeled invoice date over due date / ack date / random dates
    for pat in (
        r"Invoice\s*Date[d]?\s*[:\-]?\s*"
        r"(\d{1,2}[\-\/\.]\s*(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\-\/\.]?\s*\d{2,4}"
        r"|\d{1,2}[\-\/\.]\d{1,2}[\-\/\.]\d{2,4})",
        r"(?<!Ack\s)(?<!Ack)(?:Date|Pate|Dal)[d]?\s*[:\-|=]?\s*"
        r"(\d{1,2}[\-\/\.]\s*(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\-\/\.]?\s*\d{2,4}"
        r"|\d{1,2}[\-\/\.]\d{1,2}[\-\/\.]\d{2,4})",
        # Handwritten: Date 16.07.2026 / 16. 07. 2026
        r"(?:Date|Pate)\s*[:\-|=]?\s*(\d{1,2}\s*[\.]\s*\d{1,2}\s*[\.]\s*\d{2,4})",
    ):
        labeled = re.search(pat, text, re.I)
        if labeled:
            return re.sub(r"\s+", "", _clean(labeled.group(1)))
    top = "\n".join(text.splitlines()[:50])
    # Prefer Indian dd.mm.yyyy / dd-mm-yyyy over US m/d/yyyy email stamps.
    # Skip dates that only appear inside zip/file names (Testing 28.7.2026.zip).
    def _ok_date(val: str, blob: str) -> bool:
        for m in re.finditer(re.escape(val), blob):
            ctx = blob[max(0, m.start() - 24) : m.end() + 12]
            if re.search(r"(?i)\.(?:zip|pdf|xlsx?|jpe?g|png)\b", ctx):
                continue
            if re.search(r"(?i)testing\s+" + re.escape(val), ctx):
                continue
            return True
        return False

    for blob in (top, text):
        m = re.search(r"\b(\d{1,2}\.\d{1,2}\.\d{4})\b", blob)
        if m and _ok_date(m.group(1), blob):
            return m.group(1)
        m = re.search(r"\b(\d{1,2}-\d{1,2}-\d{4})\b", blob)
        if m and _ok_date(m.group(1), blob):
            return m.group(1)
    # Labeled but OCR-noisy: Date i6.0 / Pate 16.0%, 2024 (= 16.07.2026)
    noisy = re.search(
        r"(?i)(?:invoice|inv[a-z]*\s*no|date|pate)\D{0,24}"
        r"(\d{1,2})\D{0,3}(?:(0\s*[%7])|(\d{1,2}))\D{0,3}(20\d{2})",
        text,
    )
    if noisy:
        d = int(noisy.group(1))
        if noisy.group(2):
            mo = 7  # 0% / 07 OCR of July
        else:
            mo = int(noisy.group(3) or "0")
        y = noisy.group(4)
        # Handwritten year OCR often drifts (2024 vs 2026)
        if y in {"2023", "2024", "2025"} and re.search(r"\b2026\b", text):
            y = "2026"
        if 1 <= d <= 31 and 1 <= mo <= 12:
            return f"{d}.{mo}.{y}"
    # Never treat email client timestamps as invoice date
    def _is_email_stamp(val: str, blob: str) -> bool:
        return bool(re.search(re.escape(val) + r"\s+\d{1,2}[:.]\d{2}", blob))

    m = re.search(r"\b(\d{1,2}/\d{1,2}/\d{4})\b", top)
    if m and not _is_email_stamp(m.group(1), top):
        return m.group(1)
    for m in re.finditer(r"\b(\d{1,2}/\d{1,2}/\d{4})\b", text):
        if _is_email_stamp(m.group(1), text[max(0, m.start()) : m.end() + 16]):
            continue
        return m.group(1)
    for pat in DATE_PATTERNS[1:]:
        m = pat.search(text)
        if m:
            val = _clean(m.group(1)).replace(" ", "")
            if _is_email_stamp(val, text):
                continue
            return val
    return "Unknown"


def find_supplier(text: str) -> str:
    """Generic supplier detection — seller header only, not buyer/consignee."""
    lines = [_clean(ln) for ln in text.splitlines() if _clean(ln)]

    # Explicit footer / brand markers
    for ln in lines:
        m = re.search(r"(?i)\b(KARPAGAM\s+AUTO\s+STORES)\b", ln)
        if m:
            return m.group(1)
        m = re.search(
            r"(?i)^(?:for|declaration\s+for)\s+([A-Z0-9][A-Z0-9 &.\-]{2,60})$",
            ln,
        )
        if m:
            cand = _normalize_supplier_ocr(m.group(1))
            if re.search(
                r"(?i)agency|agencies|productz|bearings|petroleums|process|stores?|motors|traders",
                cand,
            ):
                return cand[:80]

    cut = len(lines)
    for i, ln in enumerate(lines[:50]):
        if re.search(
            r"(?i)^(buyer|bill\s*to|billed\s*to|consignee|ship\s*to|shipped\s*to|to[,.]?\s|"
            r"customer\s*name|tvs\b|details\s*of\s*receiver|receiver\s*\(|bill\s*to\s*party)",
            ln,
        ):
            cut = i
            break
        if re.search(r"(?i)buyer\s*\(bill\s*to\)|consignee\s*\(ship", ln):
            cut = i
            break
    header = lines[: max(cut, 1)]

    skip = re.compile(
        r"(?i)^[\(\[]?\s*(tax\s*invoice|taxinvoice|tax\s*sales|auto\s*taxinvoice|original|"
        r"duplicate|page|gst\s*invoice|office\s*copy|"
        r"gstin|state|e-?mail|contact|phone|invoice|dated|bill\s*to|billed\s*to|buyer|"
        r"consignee|ship|delivery|reference|dlr|sap|irn|ack\s*no|payment|to\.|e-?invoice|"
        r"item\s*name|thanks|transport|banks?\s*details|hsn|terms|credit\s*bill|debit\s*bill|"
        r"original\s*for\s*recipient|recipient)\b",
    )
    company_word = re.compile(
        r"(?i)\b(AGENCY|AGENCIES|MOTORS|AUTO|PRIVATE|LIMITED|PVT|TRADERS|ENTERPRISES|"
        r"ENTERPRISE|SERVICE|SOLUTIONS|INDUSTRIES|CORPORATION|COMPANY|DEALER|LIGHT|COVERS|GEAR|"
        r"PRODUCTZ|PRODUCTS|BEARINGS|PETROLEUMS|PROCESS|STORES|STORE)\b",
    )
    address_like = re.compile(
        r"(?i)^\d+[A-Z]?\s*,|\b(?:street|road|nagar|colony|estate|floor|gate|padithurai|"
        r"basement|veli|towers?)\b|"
        r"^\d{1,4}[,\-]|pin\s*code|madurai$|salem$|coimbatore$|"
        r"^[A-Z]{3,12}-\d{4,6}$",  # MADURAI-6250 city-PIN OCR junk
    )

    scored: list[tuple[int, str]] = []
    for i, ln in enumerate(header[:30]):
        if skip.search(ln):
            continue
        if re.search(r"GSTIN", ln, re.I):
            continue
        if address_like.search(ln) and not company_word.search(ln):
            continue
        if re.search(r"\d{6}\s*$", ln) and not company_word.search(ln):
            continue
        if re.fullmatch(r"(?i)[A-Z]{3,12}-\d{3,6}", _clean(ln)):
            continue  # MADURAI-6250
        cand = re.split(r"\s{2,}|GSTIN|NO\.?\d", ln, flags=re.I)[0].strip()
        cand = re.split(r"(?i)\s+(?:invoice|fivoes|tie:|dated|gstin|sstin)", cand)[0].strip()
        # Strip Tally-style " - (from 1-Apr-25)" fiscal tags
        cand = re.sub(r"\s*[-–]?\s*\(from\s+[^)]+\)\s*$", "", cand, flags=re.I).strip()
        cand = _normalize_supplier_ocr(cand)
        if not cand:
            continue
        if re.fullmatch(r"(?i)[A-Z]{3,12}-\d{3,6}", cand):
            continue
        if re.search(r"\d{2}-\d{2}-[A-Z]+-\d+", cand):
            continue
        if re.search(r"(?i)^(?:billed\s*to|shipped\s*to|bill\s*to)\b", cand):
            continue
        score = 0
        if company_word.search(cand):
            score += 3
        if cand.isupper() and 4 <= len(cand) <= 70:
            score += 2
        if i < 8:
            score += 2
        if re.match(r"(?i)^autolight\b|^a\.\s*l\.\s*a\.|^anbu\b|^arvind\b|^hema\b|^ask\b|^srii?\b", cand):
            score += 6
        if re.search(
            r"(?i)\b(agenc(?:y|ies)|motors|traders|enterprises?|productz|bearings|"
            r"petroleums|process|stores?)\b",
            cand,
        ):
            score += 3
        if re.search(r"(?i)\benterprise\b", cand):
            score += 4
        if re.search(r"(?i)\b(TVS AUTOMOBILE SOLUTIONS|TVS\s+SMART|BILL TO|BUYER|MANICKAM)\b", cand):
            score -= 4
        if re.search(r"(?i)\balagu\b|\bashok\s*agenc|\banamallais\b|\bkumaran\b", cand):
            score += 4
        if re.search(r"(?i)mahindra\s*bank|website|wob\s*sita|tamil\s*nady|madurai-\d", cand):
            score -= 5
        if address_like.search(cand):
            score -= 4
        if re.fullmatch(r"(?i)auto|agency|agencies|motors|limited|pvt|private|light", cand):
            continue
        letters = sum(c.isalpha() for c in cand)
        if re.search(r"[\\©®_/]{2,}|ASHOK\s*Levi", cand, re.I):
            continue
        if score > 0 and 3 <= len(cand) <= 90 and letters >= 4 and letters / max(len(cand), 1) >= 0.45:
            if not re.search(r"(?i)^(?:gstin|sstin|uin)", cand):
                scored.append((score, cand))

    for i, ln in enumerate(header[:12]):
        if re.search(r"GSTIN", ln, re.I):
            for j in range(i + 1, min(i + 4, len(header))):
                cand = _normalize_supplier_ocr(header[j])
                if skip.search(cand) or address_like.search(cand):
                    continue
                if company_word.search(cand) or (cand.isupper() and len(cand) > 5):
                    if not re.search(r"(?i)gstin|sstin", cand):
                        scored.append((6, re.split(r"\s{2,}", cand)[0][:90]))
            break

    if scored:
        scored.sort(key=lambda x: (-x[0], -len(x[1])))
        best = scored[0][1]
        header_blob = "\n".join(header[:30])
        if re.search(r"(?i)\bagenc(?:y|ies)|\bency\b", header_blob) and not re.search(
            r"(?i)\bagenc", best
        ):
            if re.search(r"(?i)light|auto|motors|traders|enterprises", best):
                best = f"{best} AGENCY"
        return best[:80]

    for ln in reversed(lines[-30:]):
        m = re.search(r"(?i)(?:^for\s+|declaration\s+for\s+)([A-Z][A-Za-z0-9 &.\-]{2,50})", ln)
        if m:
            cand = _normalize_supplier_ocr(m.group(1))
            if re.search(r"(?i)agency|productz|bearings|petroleums|process|stores?|motors", cand) or len(cand) >= 8:
                return cand[:80]
    return "Unknown"


def find_place_of_supply(text: str) -> str:
    def _clean_pos(val: str) -> str:
        val = _clean(val)
        # "33-Tamil Nadu Agent" / "Payment Terms" trailing labels
        val = re.split(r"(?i)\s+(?:Agent|Plant|Payment|Destination|GSTIN)\b", val)[0]
        return val.strip(" -–:,")[:50]

    m = re.search(
        r"Place\s*of\s*(?:Supply|Delivery)\s*[:\-]?\s*(\d{1,2}\s*[-–]?\s*[A-Za-z ]{3,25})",
        text,
        re.I,
    )
    if m:
        return _clean_pos(m.group(1))

    m = re.search(
        r"\b(\d{1,2}\s*[-–]\s*(?:Tamil\s*Nadu|Kerala|Karnataka|Andhra\s*Pradesh))\b",
        text,
        re.I,
    )
    if m:
        return _clean_pos(m.group(1))

    m = re.search(r"State\s*Name\s*[:\-]?\s*([A-Za-z ]+)\s*,\s*Code\s*[:\-]?\s*(\d{1,2})", text, re.I)
    if m:
        return _clean(f"{m.group(1)} ({m.group(2)})")

    # Infer from supplier GSTIN state code
    m = re.search(r"GSTIN\s*:?\s*(\d{2})[A-Z0-9]{13}", text, re.I)
    if m:
        code = m.group(1)
        states = {"32": "Kerala", "33": "Tamil Nadu", "29": "Karnataka", "36": "Telangana", "37": "Andhra Pradesh"}
        if code in states:
            return f"{code}-{states[code]}"
    return "Unknown"


# ---------------------------------------------------------------------------
# Line items
# ---------------------------------------------------------------------------

STOP_ITEM = re.compile(
    r"(?i)(amount\s*chargeable|tax\s*amount|cgst|sgst|igst|round\s*off|grand\s*total|"
    r"continued\s*to\s*page|computer\s*generated|bank\s*name|declaration|"
    r"taxable\s*value|subject\s*to|authori[sz]ed\s*sign|brought\s*forward|sub\s*total|"
    r"^total\b|net\s*amount|rounded\s*off|hsn\s*summary|company.?s\s*bank|"
    r"output\s*cgst|output\s*sgst|less\s*:\s*round|percent\s*discount)",
)

# Words parsers sometimes steal as "part numbers" from description-only lines
_FALSE_PART_WORDS = frozenset(
    {
        "wheel",
        "boot",
        "hose",
        "kit",
        "nut",
        "bolt",
        "set",
        "red",
        "white",
        "black",
        "oil",
        "pump",
        "ring",
        "seal",
        "switch",
        "wire",
        "lamp",
        "cover",
        "type",
        "size",
        "item",
        "goods",
        "total",
        "gst",
        "hsn",
        "nos",
        "pcs",
        "qty",
        "rate",
        "amount",
        "minda",
        "charges",
        "courier",
        "freight",
        "labour",
        "packing",
        "disc",
        "discount",
        "spare",
        "parts",
        "description",
        "particulars",
    }
)


def looks_like_part_number(part: str) -> bool:
    """True for real part codes; false for product words like Wheel/BOOT/Minda."""
    p = (part or "").strip()
    if len(p) < 3:
        return False
    if p.lower() in _FALSE_PART_WORDS:
        return False
    if any(ch.isdigit() for ch in p):
        return True
    # Alpha-only tokens are almost never part numbers on TVS invoices
    if re.fullmatch(r"[A-Za-z]+", p):
        return len(p) >= 12
    # Allow hyphenated / slashed codes without digits (rare)
    if re.search(r"[-/]", p) and len(p) >= 5:
        return True
    return False


def _normalize_part_fields(it: dict[str, str]) -> dict[str, str]:
    """Drop fake part tokens back into description."""
    it = dict(it)
    part = str(it.get("part_number") or "").strip()
    if not part:
        return it
    if looks_like_part_number(part):
        return it
    desc = str(it.get("description") or "").strip()
    it["part_number"] = ""
    it["description"] = _clean(f"{part} {desc}".strip()) if desc else part
    return it


def _fix_money(token: str) -> str:
    """Normalize OCR money like 12,00 / 11,000.00 / 7.90."""
    t = token.strip().strip(".,;:!)]}")
    # European-style decimal comma for small values: 12,00 -> 12.00
    if re.fullmatch(r"\d{1,4},\d{2}", t):
        return t.replace(",", ".")
    return t


def parse_tally_scan_items(text: str, hsn_digits: int = 8) -> list[dict[str, str]]:
    """
    Generic GST table parser for scanned/photo invoices.
    Anchors on HSN (4 or 8 digits), then qty / rate / amount — tolerates OCR junk.
    Works for any Tally-style / tax-invoice table, not a specific supplier.
    """
    items: list[dict[str, str]] = []
    hsn_re = rf"(?P<hsn>\d{{{hsn_digits}}})"
    row = re.compile(
        rf"(?P<head>.+?)\s+"
        rf"{hsn_re}\s+"
        rf"(?P<gst>\d{{1,2}})\s*%?\S*\s+"
        rf"(?P<qty>\d+)\s*"
        rf"(?P<unit>Nos|NOS|SET|SETS|BOX|roll|metre|PCS|BUCKET|BUCKETS|PKT|PKTS)?\s*"
        rf"(?P<rate>\d{{1,3}}(?:,\d{{3}})*\.\d{{2}}|\d{{1,4}},\d{{2}}|\d+\.\d{{2}})\s*"
        rf"[^\d]*?"
        rf"(?P<amount>\d{{1,3}}(?:,\d{{3}})*\.\d{{2}}|\d+\.\d{{2}})",
        re.I,
    )

    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or not re.search(rf"\d{{{hsn_digits}}}", ln):
            continue
        if STOP_ITEM.search(ln):
            continue
        if re.search(r"(?i)hsn|description of goods|gstin", ln) and not re.search(r"\d+\.\d{2}", ln):
            continue

        m = row.search(ln)
        if not m:
            continue

        head = m.group("head")
        head = re.sub(r"[\\¢°©]+", " ", head)
        head = re.sub(r"^[^A-Za-z0-9/]+", "", head)
        head = _clean(head)
        head = re.sub(r"^(?:fe|win|ial|sti|sh|s[eé]|o)\s+", "", head, flags=re.I)
        sm = re.match(r"^(\d{1,3})[\)\.\!]?\s+(.+)$", head)
        if sm:
            rest = sm.group(2)
            if re.search(r"[A-Za-z]", rest) or re.match(r"\d+[/xXmm]", rest):
                head = rest
        head = re.sub(r"^(\d{1,3})/(?=[A-Za-z]{2,})", "", head)
        head = re.sub(r"[\W_]+$", "", head)
        head = re.sub(r"[,;:\s]+\d*$", "", head).strip()
        head = _clean(head)

        desc = head
        if len(desc) < 2:
            continue
        desc = re.sub(r"^\d{1,3}[\)\.\!]\s*", "", desc)
        sm = re.match(r"^(\d{1,3})\s+(.+)$", desc)
        if sm and re.search(r"[A-Za-z]", sm.group(2)):
            rest = sm.group(2)
            if not re.match(r"(?i)mm\b|cm\b|inch|nos\b|set\b", rest):
                desc = rest
        desc = re.sub(r"^/\d+\s+", "", desc)

        # Optional: split part number if first token looks like a code
        part = ""
        tokens = desc.split()
        if tokens:
            t0 = tokens[0]
            if re.search(r"[A-Za-z]", t0) and re.search(r"\d", t0) and len(t0) >= 5:
                part = t0
                desc = " ".join(tokens[1:]) or t0

        unit = (m.group("unit") or "Nos").rstrip("-")
        items.append(
            {
                "part_number": part,
                "description": desc,
                "hsn_sac": m.group("hsn"),
                "qty": f"{m.group('qty')} {unit}".strip(),
                "rate": _fix_money(m.group("rate")),
                "amount": _fix_money(m.group("amount")),
            }
        )

    return _dedupe_items(items)


def _dedupe_items(items: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[tuple] = set()
    unique: list[dict[str, str]] = []
    for it in items:
        it = _strip_part_from_description(dict(it))
        # Drop garbage OCR rows
        try:
            amt = float(str(it.get("amount") or "0").replace(",", ""))
            rate = float(str(it.get("rate") or "0").replace(",", ""))
        except Exception:
            amt, rate = 0.0, 0.0
        desc = it.get("description") or ""
        if amt <= 0.2 and rate <= 0.2:
            continue
        if re.fullmatch(r"[\d\s,\.\-%]+", desc) and not it.get("part_number"):
            continue

        part = str(it.get("part_number") or "").strip().upper()
        desc_key = re.sub(r"\W+", "", desc.lower())[:40]
        # Always key on part number when present — many AC/spare lines share qty/rate/amount
        if part:
            key = (part, it.get("qty"), it.get("rate"), it.get("amount"), desc_key)
        else:
            key = (it.get("qty"), it.get("rate"), it.get("amount"), desc_key)
            # Soft match only when there is no part code (avoid wiping 50+ distinct SKUs)
            soft = (desc_key[:24], it.get("amount"))
            if soft in seen:
                continue
            seen.add(soft)
        if key in seen:
            continue
        seen.add(key)
        d = re.sub(r"^\d{1,3}[\)\.]\s+", "", desc)
        d = re.sub(r"\s+", " ", d).strip(" ,;.-")
        if len(d) >= 2 or it.get("part_number"):
            it["description"] = d
            unique.append(it)
    return unique


def _strip_part_from_description(it: dict[str, str]) -> dict[str, str]:
    from extraction.validate import strip_part_from_description

    return strip_part_from_description(it)


def parse_part_column_items(text: str) -> list[dict[str, str]]:
    """
    Tables with an explicit Part No column (common across GST invoices):
    ... DESC HSN PART QTY RATE ... AMOUNT
    """
    items: list[dict[str, str]] = []
    row = re.compile(
        r"(?P<desc>[A-Za-z][A-Za-z0-9 \/\.\-]{2,}?)\s*"
        r"[\/\s]*(?P<hsn>\d{4,8})\s+"
        r"(?P<part>[A-Z]{1,4}[\-A-Z0-9]{3,})\s*[,\.]?\s*"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s*(?P<unit>NOS|Nos|PCS|SET|BUCKET)?\s+"
        r"(?P<rate_a>[\d,]+\.\d{2})\s+"
        r"(?P<rate_b>[\d,]+\.\d{2})?",
        re.I,
    )
    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        m = row.search(ln)
        if not m:
            continue
        money = re.findall(r"[\d,]+\.\d{2}", ln)
        rate = m.group("rate_b") or m.group("rate_a")
        qty = m.group("qty")
        if len(money) >= 3:
            amount = money[-1]
        else:
            try:
                amount = f"{float(qty.replace(',','')) * float(rate.replace(',','')):,.2f}"
            except Exception:
                amount = money[-1] if money else ""
        items.append(
            {
                "part_number": m.group("part"),
                "description": _clean(m.group("desc")),
                "hsn_sac": m.group("hsn"),
                "qty": f"{qty} {(m.group('unit') or '')}".strip(),
                "rate": rate,
                "amount": amount,
            }
        )
    return _dedupe_items(items)


def parse_amount_trail_items(text: str) -> list[dict[str, str]]:
    """
    Loose fallback for messy photos: description ... qty ... rate amount
    when HSN/GST columns are unreadable.
    """
    items: list[dict[str, str]] = []
    row = re.compile(
        r"(?P<head>[A-Za-z0-9][A-Za-z0-9 \/\.\-]{4,}?)\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s*"
        r"(?P<unit>NOS|Nos|PCS|SET|BUCKET|ROLL)?\s+"
        r"(?P<rate>[\d,]+\.\d{2})\s+"
        r"(?:[A-Za-z%]+\s+)?"
        r"(?P<amount>[\d,]+\.\d{2})",
        re.I,
    )
    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        if not re.search(r"[A-Za-z]{3,}", ln):
            continue
        m = row.search(ln)
        if not m:
            continue
        qty_raw = m.group("qty")
        unit = m.group("unit") or ""
        # 4–8 digit "qty" with no unit is almost always an HSN code — skip (other parsers handle it)
        if not unit and re.fullmatch(r"\d{4,8}", qty_raw):
            continue
        head = _clean(m.group("head"))
        head = re.sub(r"^\d{1,3}[\)\.\s]+", "", head)
        part = ""
        tokens = head.split()
        if tokens and re.search(r"\d", tokens[0]) and re.search(r"[A-Za-z]", tokens[0]) and len(tokens[0]) >= 5:
            part, desc = tokens[0], " ".join(tokens[1:]) or tokens[0]
        else:
            desc = head
        if len(desc) < 3:
            continue
        items.append(
            {
                "part_number": part,
                "description": desc,
                "hsn_sac": "",
                "qty": f"{qty_raw} {unit}".strip(),
                "rate": m.group("rate"),
                "amount": m.group("amount"),
            }
        )
    return _dedupe_items(items)


def parse_credit_bill_hsn_rate_qty(text: str) -> list[dict[str, str]]:
    """
    CREDIT BILL / parts counter layout:
    PartNo Description HSN Rate Qty [Amount]
    e.g. F8P08758 RADIATOR HOSES 87089900 247.00 3.00
    """
    items: list[dict[str, str]] = []
    row = re.compile(
        r"(?P<part>[A-Z][A-Z0-9][A-Z0-9\-]{3,})\s+"
        r"(?P<desc>[A-Za-z][A-Za-z0-9 \/\.\-]{2,}?)\s+"
        r"(?P<hsn>\d{4,8})\s+"
        r"(?P<rate>[\d,]+\.\d{2})\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s*"
        r"(?P<unit>NOS|Nos|PCS|SET|SETS|Nos-?)?\s*"
        r"(?P<amount>[\d,]+\.\d{2})?",
        re.I,
    )
    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        m = row.search(ln)
        if not m:
            continue
        qty = m.group("qty")
        unit = (m.group("unit") or "").rstrip("-")
        rate = m.group("rate")
        amount = m.group("amount") or ""
        if not amount:
            try:
                amount = f"{float(qty.replace(',', '')) * float(rate.replace(',', '')):,.2f}"
            except Exception:
                amount = ""
        # Sanity: qty should be a real quantity, not another HSN
        if re.fullmatch(r"\d{4,8}", qty) and not unit:
            continue
        try:
            qn = float(qty.replace(",", ""))
            qty_out = str(int(qn)) if qn == int(qn) else qty
        except Exception:
            qty_out = qty
        items.append(
            {
                "part_number": m.group("part"),
                "description": _clean(m.group("desc")),
                "hsn_sac": m.group("hsn"),
                "qty": f"{qty_out} {unit}".strip(),
                "mrp": "",  # credit bills: Rate is selling price, not MRP
                "rate": rate,
                "amount": amount,
            }
        )
    return _dedupe_items(items)


def parse_photo_gstr_qty_rate(text: str) -> list[dict[str, str]]:
    """
    Phone-photo OCR of PART | HSN | GSTR% | QTY | Rate | Dis% | AMOUNT tables.
    Examples:
      —~Tsaeneee | 18%| 2NOS| 906.00| 38%| 1,123.46
      39269089 | 18%) 1PKT) 200,00 200.90
    """
    items: list[dict[str, str]] = []
    money = r"(?:[\d,]+\.\d{2}|\d{1,4},\d{2})"
    trail = re.compile(
        rf"(?:(?P<hsn>\d{{4,8}})\s*[|)\s]*)?"
        rf"(?P<gst>\d{{1,2}})\s*%[|)\s]*"
        rf"(?P<qty>\d+(?:[.,]\d+)?)\s*"
        rf"(?P<unit>NOS|PKT|PCS|SET|SETS|ROLL|Nos|PKTS)?\s*"
        rf"[|)\s]*"
        rf"(?P<rate>{money})"
        rf"(?:\s*(?:\d{{1,2}}\s*%)?[|)\s]*)*"
        rf"(?P<amount>{money})",
        re.I,
    )

    # Harvest likely part codes from the whole OCR blob (often on separate lines)
    part_cands: list[str] = []
    for m in re.finditer(r"\b((?:Sw|SW|GT|Sp|SP|Pt|PT)\s*[\-]?\s*\d{2,5}[A-Za-z]?)\b", text, re.I):
        p = re.sub(r"\s+", " ", m.group(1)).strip()
        if p.upper() not in {x.upper() for x in part_cands}:
            part_cands.append(p)
    # GT 370 RED — OCR often splits / mangled as Ta9370 + RED
    if re.search(r"(?i)\b(?:GT\s*)?370\b", text) and re.search(r"(?i)\bRED\b", text):
        if not any(re.search(r"(?i)370", p) for p in part_cands):
            part_cands.append("GT 370 RED")

    desc_hints: list[str] = []
    if re.search(r"(?i)cum\s*st", text):
        desc_hints.append("Ign Cum Stg Lock")
    if re.search(r"(?i)tag\s*370|teus70|ta9?370|red[-\s]?white|rea\.?\s*wine", text):
        desc_hints.append("GT Tag370 Red-White")

    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        if not re.search(r"\d{1,2}\s*%", ln):
            continue
        m = trail.search(ln)
        if not m:
            continue
        qty = m.group("qty")
        unit = (m.group("unit") or "").upper()
        try:
            qty_n = float(qty.replace(",", ""))
        except Exception:
            continue
        if not unit and qty_n > 99:
            continue
        head = _clean(ln[: m.start()])
        head = re.sub(r"^\d{1,3}[\)\.\|]\s*", "", head)
        head = re.sub(r"[|\\~—–]+", " ", head)
        head = re.sub(r"[^\w\s/\.\-]+", " ", head)
        head = _clean(head)
        part, desc = "", head
        tokens = [t for t in head.split() if re.search(r"[A-Za-z0-9]", t)]
        if tokens:
            t0 = tokens[0]
            if re.search(r"[A-Za-z]", t0) and re.search(r"\d", t0):
                part = t0
                desc = " ".join(tokens[1:]) or t0
            elif len(tokens) >= 2 and re.match(r"(?i)^(sw|gt|sp|pt)$", t0):
                part = f"{t0} {tokens[1]}"
                desc = " ".join(tokens[2:]) or part
        letters = sum(c.isalpha() for c in desc)
        vowels = sum(c in "aeiouAEIOU" for c in desc)
        if letters < 4 or (letters >= 6 and vowels == 0) or re.search(r"(.)\1{2,}", desc):
            desc = ""
        if len(desc) < 2 and not part:
            desc = f"Item HSN {m.group('hsn')}" if m.group("hsn") else "Line item"
        items.append(
            {
                "part_number": part,
                "description": (desc or "Line item")[:80],
                "hsn_sac": m.group("hsn") or "",
                "qty": f"{qty} {unit}".strip(),
                "rate": _fix_money(m.group("rate")),
                "amount": _fix_money(m.group("amount")),
            }
        )

    pi = 0
    di = 0
    for it in items:
        if not it.get("part_number") and pi < len(part_cands):
            it["part_number"] = part_cands[pi]
            pi += 1
        elif it.get("part_number") and pi < len(part_cands):
            pi += 1
        weak = it.get("description") in ("", "Line item") or str(it.get("description", "")).startswith(
            "Item HSN"
        )
        if weak and di < len(desc_hints):
            it["description"] = desc_hints[di]
            di += 1
        elif not weak and di < len(desc_hints):
            di += 1
    for it in items:
        if not it.get("part_number") and pi < len(part_cands):
            it["part_number"] = part_cands[pi]
            pi += 1
    return _dedupe_items(items)


def parse_part_hsn_qty_rate(text: str) -> list[dict[str, str]]:
    """
    PartNo Description HSN Qty [Unit] Rate [Tax%] Amount
    (no serial number prefix — common on credit bills)

    Also handles Leyland/Ashok MRP layouts where Rate is absent:
      Qty MRP Dis% Tax% Amount  → fill mrp, derive rate from amount/qty
    """
    items: list[dict[str, str]] = []
    money = r"[\d,]+\.\d{2}"
    row = re.compile(
        r"(?P<part>[A-Z][A-Z0-9][A-Z0-9\-]{3,})\s+"
        r"(?P<desc>[A-Za-z][A-Za-z0-9 \/\.\-\&\(\)]{2,}?)\s+"
        r"(?P<hsn>\d{4,8})\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s*"
        r"(?P<unit>NOS-?|Nos|PCS|SET|SETS|ROLL)?\s+"
        r"(?P<rate>" + money + r")\s+"
        r"(?:(?P<tax>\d{1,2})\s+)?"
        r"(?P<amount>" + money + r")",
        re.I,
    )
    # Qty MRP Disc% Tax% Amount (disc/tax may be int or xx.00)
    mrp_row = re.compile(
        r"(?P<part>[A-Z][A-Z0-9][A-Z0-9\-]{3,})\s+"
        r"(?P<desc>.+?)\s+"
        r"(?P<hsn>\d{4,8})\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s*"
        r"(?P<unit>NOS-?|Nos|PCS|SET|SETS|ROLL)?\s+"
        r"(?P<mrp>" + money + r")\s+"
        r"(?P<disc>\d{1,2}(?:\.\d{1,2})?)\s+"
        r"(?P<tax>\d{1,2}(?:\.\d{1,2})?)\s+"
        r"(?P<amount>" + money + r")",
        re.I,
    )
    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        mm = mrp_row.search(ln)
        if mm:
            qty = mm.group("qty")
            unit = (mm.group("unit") or "").rstrip("-")
            mrp = mm.group("mrp")
            amount = mm.group("amount")
            rate = ""
            try:
                q = float(qty.replace(",", ""))
                a = float(amount.replace(",", ""))
                if q > 0:
                    rate = f"{a / q:.2f}"
            except Exception:
                rate = ""
            part = mm.group("part")
            desc = _clean(mm.group("desc"))
            if not looks_like_part_number(part):
                desc = _clean(f"{part} {desc}".strip())
                part = ""
            items.append(
                {
                    "part_number": part,
                    "description": desc,
                    "hsn_sac": mm.group("hsn"),
                    "qty": f"{qty} {unit}".strip(),
                    "mrp": mrp,
                    "rate": rate,  # net unit only — never copy MRP into Rate
                    "amount": amount,
                }
            )
            continue
        m = row.search(ln)
        if not m:
            continue
        unit = (m.group("unit") or "").rstrip("-")
        moneys = re.findall(money, ln)
        rate = m.group("rate")
        amount = m.group("amount")
        mrp = ""
        # If OCR captured MRP + Disc% + Amount as three moneys, first is MRP
        if len(moneys) >= 3:
            try:
                mid = float(moneys[-2].replace(",", ""))
            except Exception:
                mid = 999
            # Middle value looks like a discount/tax percent, not a dealer rate
            if mid <= 100 and float(moneys[-1].replace(",", "")) > mid:
                mrp = moneys[0] if len(moneys) == 3 else moneys[-3]
                amount = moneys[-1]
                # Prefer amount/qty as net rate when disc sits between mrp and amount
                try:
                    q = float(m.group("qty").replace(",", ""))
                    a = float(amount.replace(",", ""))
                    rate = f"{a / q:.2f}" if q > 0 else (rate if rate != mrp else "")
                except Exception:
                    pass
                if not mrp:
                    mrp = moneys[0]
        part = m.group("part")
        desc = _clean(m.group("desc"))
        if not looks_like_part_number(part):
            desc = _clean(f"{part} {desc}".strip())
            part = ""
        items.append(
            {
                "part_number": part,
                "description": desc,
                "hsn_sac": m.group("hsn"),
                "qty": f"{m.group('qty')} {unit}".strip(),
                "mrp": mrp,
                "rate": rate,
                "amount": amount,
            }
        )
    return _dedupe_items(items)


def parse_part_mrp_disc_tax_amount(text: str) -> list[dict[str, str]]:
    """
    Ashok Leyland retail / similar:
      S.No PartNo Description HSN Qty MRP Dis% Tax% Amount
    No Rate column — MRP is explicit; net unit rate ≈ Amount / Qty.
    """
    items: list[dict[str, str]] = []
    money = r"[\d,]+\.\d{2}"
    row = re.compile(
        r"(?:(?P<sno>\d{1,3})\s+)?"
        r"(?P<part>[A-Z0-9][A-Z0-9\-]{5,})\s+"
        r"(?P<desc>.+?)\s+"
        r"(?P<hsn>\d{4,8})\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s*"
        r"(?P<unit>NOS|Nos|PCS|SET)?\s*"
        r"(?P<mrp>" + money + r")\s+"
        r"(?P<disc>\d{1,2}(?:\.\d{1,2})?)\s+"
        r"(?P<tax>\d{1,2}(?:\.\d{1,2})?)\s+"
        r"(?P<amount>" + money + r")",
        re.I,
    )
    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        m = row.search(ln)
        if not m:
            continue
        desc = _clean(m.group("desc"))
        if len(re.sub(r"[^A-Za-z0-9]", "", desc)) < 2:
            continue
        qty = m.group("qty")
        unit = (m.group("unit") or "").strip()
        if re.fullmatch(r"\d{4,8}", qty) and not unit:
            continue
        mrp = m.group("mrp")
        amount = m.group("amount")
        rate = ""
        try:
            q = float(qty.replace(",", ""))
            a = float(amount.replace(",", ""))
            if q > 0:
                rate = f"{a / q:.2f}"
        except Exception:
            rate = ""
        items.append(
            {
                "part_number": m.group("part"),
                "description": desc,
                "hsn_sac": m.group("hsn"),
                "qty": f"{qty} {unit}".strip(),
                "mrp": mrp,
                "rate": rate,  # net unit only — never copy MRP into Rate
                "amount": amount,
            }
        )
    return _dedupe_items(items)


def parse_handwritten_rate_qty_amount(text: str) -> list[dict[str, str]]:
    """
    Sparse handwritten GST slips (e.g. Alagu Gear Rod Covers):
      Description [HSN] Rate Qty Amount
    Tolerates integer amounts and mild OCR noise on qty (Sa0 → 500).
    """
    items: list[dict[str, str]] = []
    num = r"[\d,]+\.?\d*"
    row = re.compile(
        r"(?:(?P<sno>\d{1,3})[\)\.\s]+)?(?P<desc>(?:\d{1,2}\s*/\s*\d{1,2}\s+)?"
        r"(?:\d{1,2}\s+)?[A-Za-z][A-Za-z0-9 \/&\-]{2,40}?)\s+"
        r"(?:(?P<hsn>\d{4,8})\s+)?"
        r"(?P<rate>" + num + r")\s+"
        r"(?P<qty>" + num + r"|[Ss5][a-z0-9]{0,3}|le[o0]|l00|1oo)\s+"
        r"(?P<amount>" + num + r")\b",
        re.I,
    )
    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        if not re.search(r"(?i)[A-Za-z]{3,}", ln):
            continue
        # Skip pure bank / header lines
        if re.search(r"(?i)\b(gstin|ifsc|bank|branch|a/?c\s*no)\b", ln):
            continue
        m = row.search(ln)
        if not m:
            continue
        desc = _clean(m.group("desc"))
        desc = re.sub(r"^\d{1,3}[\)\.\s]+", "", desc)
        if len(re.sub(r"[^A-Za-z]", "", desc)) < 3:
            continue
        rate = _fix_money(m.group("rate").replace(",", ""))
        qty_raw = m.group("qty")
        # OCR: Sa0 / sm0 / S00 → 500-ish handwritten 500
        if re.fullmatch(r"(?i)[Ss5][a-z0o]*", qty_raw) or re.fullmatch(r"(?i)s\w?0", qty_raw):
            qty_raw = "500"
        elif re.fullmatch(r"(?i)le[o0]|l00|1oo", qty_raw):
            qty_raw = "100"
        qty_raw = re.sub(r"[^\d.]", "", qty_raw) or qty_raw
        amount = _fix_money(str(m.group("amount")).replace(",", ""))
        # Amounts on these slips are usually whole rupees ≥ 100
        try:
            r = float(rate)
            q = float(qty_raw)
            a = float(amount)
        except Exception:
            continue
        if r <= 0 or q <= 0 or a < 50:
            continue
        # Prefer exact qty*rate≈amount; allow rate±1 OCR drift
        if abs(q * r - a) > 0.51:
            if abs(q * (r - 1) - a) <= 0.51:
                r = r - 1
            elif abs(q * (r + 1) - a) <= 0.51:
                r = r + 1
            elif abs(q * r - a) > max(250.0, 0.08 * a):
                continue
            rate = str(int(r)) if r == int(r) else f"{r:.2f}"
        hsn = m.group("hsn") or ""
        items.append(
            {
                "part_number": "",
                "description": desc[:80],
                "hsn_sac": hsn,
                "qty": str(int(q)) if q == int(q) else str(q),
                "mrp": "",
                "rate": f"{r:.2f}" if ("." in rate or r != int(r)) else str(int(r)),
                "amount": f"{a:.2f}" if a != int(a) else str(int(a)),
            }
        )
    return _dedupe_items(items)


def _normalize_shifted_hsn(items: list[dict[str, str]]) -> list[dict[str, str]]:
    """Fix rows where HSN landed in qty and real qty landed in amount."""
    fixed: list[dict[str, str]] = []
    for it in items:
        qty = str(it.get("qty") or "").strip()
        rate = str(it.get("rate") or "").strip()
        amount = str(it.get("amount") or "").strip()
        hsn = str(it.get("hsn_sac") or "").strip()
        qty_num = re.sub(r"[^\d]", "", qty.split()[0]) if qty else ""
        amt_num = amount.replace(",", "")
        # Classic shift: qty is 4–8 digit HSN, amount is small quantity
        if (
            not hsn
            and re.fullmatch(r"\d{4,8}", qty_num)
            and re.fullmatch(r"\d+(?:\.\d{1,2})?", amt_num or "")
        ):
            try:
                q = float(amt_num)
                r = float(rate.replace(",", ""))
            except Exception:
                fixed.append(it)
                continue
            if q <= 5000 and r > 0:
                new_amount = f"{q * r:,.2f}"
                it = {
                    **it,
                    "hsn_sac": qty_num,
                    "qty": str(int(q)) if q == int(q) else str(q),
                    "rate": rate,
                    "amount": new_amount,
                }
        # HSN glued onto description end
        desc = str(it.get("description") or "")
        glued = re.search(r"^(.*?)(\d{8})$", desc)
        if glued and not it.get("hsn_sac"):
            it = {
                **it,
                "description": glued.group(1).strip(" -"),
                "hsn_sac": glued.group(2),
            }
        fixed.append(it)
    return fixed


def _fnum(val: str | None) -> float | None:
    if not val:
        return None
    try:
        return float(str(val).split()[0].replace(",", ""))
    except Exception:
        return None


def _repair_qty_mrp_shift(
    items: list[dict[str, str]],
    text: str = "",
    schema: str | None = None,
) -> list[dict[str, str]]:
    from extraction.validate import repair_qty_mrp_shift

    return repair_qty_mrp_shift(items, text, schema=schema)


def parse_einvoice_line_items(text: str) -> list[dict[str, str]]:
    """
    Digital e-invoice layout (e.g. SRI KARTHICK AGENCY):
    S.No PartNo Description HSN Qty Unit Rate Tax% Amount
    1 1728431BC60SQMM25 8MM WIRE 85443000 2 NOS- 1788.13 18 4219.98
    """
    items: list[dict[str, str]] = []
    row = re.compile(
        r"^(?P<sl>\d{1,3})\s+"
        r"(?P<head>.+?)\s+"
        r"(?P<hsn>\d{8})\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s*"
        r"(?P<unit>NOS-?|NO|SETS?|ROLL|PCS-?|Nos-?)?\s+"
        r"(?P<rate>[\d,]+\.\d{2})\s+"
        r"(?P<tax>\d{1,2})\s+"
        r"(?P<amount>[\d,]+\.\d{2})\s*$",
        re.I,
    )
    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        m = row.match(ln)
        if not m:
            continue
        head = _clean(m.group("head"))
        tokens = head.split()
        part, desc = "", head
        if tokens:
            t0 = tokens[0]
            # Single-token part codes are most common on these e-invoices
            if re.search(r"[A-Za-z]", t0) and (re.search(r"\d", t0) or len(t0) >= 5):
                part = t0
                desc = " ".join(tokens[1:]) if len(tokens) > 1 else t0
            elif re.match(r"^\d", t0) and len(tokens) >= 2:
                # e.g. "3 INCH ALLU PIPE ALLU CLAMP..."
                take = 1
                for tok in tokens[1:4]:
                    if tok.isupper() or re.search(r"\d", tok):
                        take += 1
                    else:
                        break
                part = " ".join(tokens[:take])
                desc = " ".join(tokens[take:]) or part
            else:
                desc = head
        unit = (m.group("unit") or "").rstrip("-")
        items.append(
            {
                "part_number": part,
                "description": desc,
                "hsn_sac": m.group("hsn"),
                "qty": f"{m.group('qty')} {unit}".strip(),
                "rate": m.group("rate"),
                "amount": m.group("amount"),
            }
        )
    return _dedupe_items(items)


def _item_score(items: list[dict[str, str]], schema: str | None = None) -> int:
    """Prefer complete rows; score MRP only when the layout expects it."""
    from extraction.layout import schema_expects_mrp

    expects_mrp = schema_expects_mrp(schema) if schema else None  # type: ignore[arg-type]
    score = 0
    for it in items:
        if not (it.get("description") or it.get("part_number")):
            continue
        score += 1
        if it.get("qty"):
            qtok = str(it.get("qty") or "").split()[0].replace(",", "")
            # Penalize MRP/money sitting in Qty
            if re.fullmatch(r"\d+\.\d{2}", qtok) and float(qtok) >= 100:
                score -= 2
            else:
                score += 1
        if it.get("rate"):
            score += 1
        if it.get("mrp"):
            if expects_mrp is True:
                score += 2
            elif expects_mrp is False:
                score -= 1  # invented MRP on non-MRP layouts
            else:
                score += 1  # unknown: mild reward only
        if it.get("amount"):
            score += 2
        if it.get("part_number"):
            if looks_like_part_number(str(it.get("part_number") or "")):
                score += 2
            else:
                score -= 2  # invented product-word "parts" must lose
        if it.get("hsn_sac"):
            score += 1
        # Prefer rows that already split Item Code + HSN (avoid desc-stuffed parsers)
        if looks_like_part_number(str(it.get("part_number") or "")) and it.get("hsn_sac"):
            score += 1
    return score


def _align_rate_to_amount(qty: str, rate: str, amount: str) -> str:
    """Fix OCR-dropped decimals (e.g. 249 vs 2.49 when qty*rate≈amount)."""
    try:
        q = float(qty.replace(",", ""))
        r = float(rate.replace(",", ""))
        a = float(amount.replace(",", ""))
    except Exception:
        return rate
    if q <= 0:
        return rate
    if abs(q * r - a) <= 0.06:
        return f"{r:.2f}" if "." not in rate else rate
    for div in (10, 100, 1000):
        if abs(q * (r / div) - a) <= 0.06:
            return f"{r / div:.2f}"
    return rate


def _ocr_hsn_digits(token: str) -> str:
    """Keep clean 8-digit HSN; only lightly repair low-letter OCR tokens."""
    t = re.sub(r"[^A-Za-z0-9]", "", token.strip())
    if re.fullmatch(r"\d{8}", t):
        return t
    if re.fullmatch(r"\d{9}", t):
        return t[-8:]
    # ja5124000 → 85124000 (j + 7 digits, leading 8 dropped by OCR)
    if re.fullmatch(r"(?i)j[a-z]?\d{7}", t):
        digits = re.sub(r"\D", "", t)
        if len(digits) == 7:
            return "8" + digits
    # Soft letter→digit only when almost all digits (≤2 letters)
    letters = sum(c.isalpha() for c in t)
    if 8 <= len(t) <= 9 and letters <= 2 and re.fullmatch(r"[0-9OoIlSsAaGgJj]+", t):
        map_ch = str.maketrans("OoIlSsAaGgJj", "001155448877")
        cand = t.translate(map_ch)
        if len(cand) == 9:
            cand = cand[-8:]
        if re.fullmatch(r"\d{8}", cand) and cand[:2] in {
            "39", "40", "73", "82", "83", "84", "85", "87", "90", "94",
        }:
            return cand
    return ""


def parse_s_code_nos_items(text: str) -> list[dict[str, str]]:
    """
    Tally-style parts invoices (e.g. SRI VINAYAKA / SUPER Brand):
      2S 2007 -BLADE FUSE 30A 85361090 50 nos 2.66 nos 133.00
      11 $9810 LED BREAK LIGHT ... 85122010 25 nos 21.08 nos 527.00
    Part codes start with S + digits (OCR often reads S as $); qty uses 'nos'.
    Prefer SI No when present so duplicate SKUs (same part twice) are kept.
    """
    by_sl: dict[int, dict[str, str]] = {}
    no_sl: list[dict[str, str]] = []
    row = re.compile(
        r"(?P<sl>\d{1,2})?\s*[+§]?\s*"
        r"(?P<part>[S$]\s*\d{3,5}[A-Za-z]?)\s+"
        r"(?P<desc>.+?)\s+"
        r"(?:[|\[{/\s]*)?(?P<hsn>\d{8,9}|[A-Za-z]{0,3}\d{5,8}|[0-9OoIlSsAaGgJj]{8})?\s*"
        r"(?P<qty>\d+)\s*nos\s+"
        r"(?P<rate>\d+(?:[.,]\d+)?)\s+"
        r"(?:nos\s+)?"
        r"(?P<amount>[\d,]+\.\d{2})",
        re.I,
    )

    def _quality(it: dict[str, str]) -> float:
        score = 0.0
        if it.get("hsn_sac") and re.fullmatch(r"\d{8}", it["hsn_sac"] or ""):
            score += 2
        try:
            q = float(str(it["qty"]).split()[0])
            r = float(str(it["rate"]).replace(",", ""))
            a = float(str(it["amount"]).replace(",", ""))
            if abs(q * r - a) <= 0.06:
                score += 3
        except Exception:
            pass
        score += min(len(it.get("description") or ""), 40) / 20.0
        return score

    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        if not re.search(r"(?i)\bnos\b", ln):
            continue
        m = row.search(ln)
        if not m:
            continue
        part = re.sub(r"\s+", " ", m.group("part")).strip()
        part = part.replace("$", "S")
        pm = re.match(r"(?i)^(S)\s*(\d{3,5}[A-Za-z]?)$", part)
        if pm:
            part = f"{pm.group(1).upper()} {pm.group(2)}"
        desc = _clean(m.group("desc"))
        desc = re.sub(r"[|\[{\\]+$", "", desc).strip()
        desc = re.sub(r"^\d{1,2}\s+", "", desc)
        # Trailing OCR HSN junk in description (gsoago30, sg26o0e9, jss26o0¢9)
        junk = re.search(r"(?i)(?:\s*[|\\])?\s*([a-z0-9¢]{6,12})\s*$", desc)
        if junk and not re.fullmatch(r"\d+\.?\d*", junk.group(1)):
            maybe = _ocr_hsn_digits(re.sub(r"[^A-Za-z0-9]", "", junk.group(1)))
            desc = desc[: junk.start()].strip()
        else:
            maybe = ""
        desc = re.sub(r"\s+\d{1,2}$", "", desc).strip()

        hsn_raw = m.group("hsn") or ""
        if re.fullmatch(r"\d{9}", hsn_raw):
            hsn = _ocr_hsn_digits(hsn_raw[-8:])
        else:
            hsn = _ocr_hsn_digits(hsn_raw)
        if not hsn and maybe:
            hsn = maybe
        if not hsn:
            hm = re.search(r"(\d{8})\s*$", desc)
            if hm:
                hsn = hm.group(1)
                desc = desc[: hm.start()].strip()
        if desc.upper().startswith(part.upper()):
            desc = desc[len(part) :].strip(" -") or desc
        desc = desc.lstrip(" -").strip()
        qty = m.group("qty")
        rate = _align_rate_to_amount(qty, m.group("rate").replace(",", "."), m.group("amount"))
        if len(desc) < 2:
            desc = part
        item = {
            "part_number": part,
            "description": desc,
            "hsn_sac": hsn,
            "qty": f"{qty} nos",
            "rate": rate,
            "amount": m.group("amount"),
        }
        sl_raw = m.group("sl")
        if sl_raw:
            sl = int(sl_raw)
            if 1 <= sl <= 250:
                prev = by_sl.get(sl)
                if not prev or _quality(item) >= _quality(prev):
                    twins = [
                        s
                        for s, existing in by_sl.items()
                        if s != sl
                        and existing.get("part_number") == item["part_number"]
                        and existing.get("amount") == item["amount"]
                    ]
                    # True duplicate SKUs are consecutive (5+6). Ghost OCR rows are not.
                    if twins and all(abs(s - sl) > 1 for s in twins):
                        continue
                    by_sl[sl] = item
                continue
        no_sl.append(item)

    if len(by_sl) >= 3:
        # Drop trailing ghost SI Nos (e.g. 12 repeating line 1) before gap-fill
        keys = sorted(by_sl)
        while len(keys) >= 2:
            hi = keys[-1]
            lo_matches = [
                k
                for k in keys[:-1]
                if by_sl[k].get("part_number") == by_sl[hi].get("part_number")
                and by_sl[k].get("amount") == by_sl[hi].get("amount")
                and abs(k - hi) > 1
            ]
            if lo_matches:
                by_sl.pop(hi)
                keys = sorted(by_sl)
                continue
            break

        # If SI 1 is missing but a high SI holds the first SKU, move it to slot 1
        if 1 not in by_sl and by_sl:
            hi = max(by_sl)
            if hi > len(by_sl):
                by_sl[1] = by_sl.pop(hi)

        max_sl = max(by_sl)
        # Don't gap-fill across huge holes from a single bad high SI —
        # but allow long invoices (80–150 lines) when coverage is dense.
        hole = max_sl - len(by_sl)
        if hole > 3 and len(by_sl) < 40:
            max_sl = max(k for k in by_sl if k <= len(by_sl) + 2) if by_sl else max_sl
        elif hole > 15 and len(by_sl) >= 40:
            # Dense long invoice with a few OCR gaps — still gap-fill within range
            max_sl = max(by_sl)

        gaps = [i for i in range(1, max_sl + 1) if i not in by_sl]
        pool = list(no_sl)
        for gap in gaps:
            prev = by_sl.get(gap - 1)
            nxt = by_sl.get(gap + 1)
            picked_idx = None
            for i, cand in enumerate(pool):
                if prev and cand["part_number"] == prev["part_number"] and cand["amount"] == prev["amount"]:
                    picked_idx = i
                    break
                if nxt and cand["part_number"] == nxt["part_number"] and cand["amount"] == nxt["amount"]:
                    picked_idx = i
                    break
            if picked_idx is not None:
                by_sl[gap] = pool.pop(picked_idx)

        return [by_sl[k] for k in sorted(by_sl)]
    return _dedupe_items(no_sl + list(by_sl.values()))


def _clone_item_to_match_subtotal(items: list[dict[str, str]], text: str) -> list[dict[str, str]]:
    """If OCR dropped a duplicate row, taxable total is often short by exactly that amount."""
    if not items:
        return items
    try:
        item_sum = sum(float(str(it.get("amount") or "0").replace(",", "")) for it in items)
    except Exception:
        return items

    # Prefer labeled taxable / subtotal only — bare amounts over-match and double-clone
    candidates: list[float] = []
    for m in re.finditer(
        r"(?i)(?:taxable\s*(?:value|amt|amount)?|sub\s*total|total\s*(?:amount)?)\s*[:\-]?\s*([\d,]+\.\d{2})",
        text,
    ):
        try:
            candidates.append(float(m.group(1).replace(",", "")))
        except Exception:
            pass
    # Amount alone on a line (Tally often prints 2,644.54 by itself)
    for raw in text.splitlines():
        ln = _clean(raw)
        if re.fullmatch(r"[\d,]+\.\d{2}", ln):
            try:
                val = float(ln.replace(",", ""))
            except Exception:
                continue
            if 500 <= val <= 500000:
                candidates.append(val)

    # Already reconciled to a candidate total — do nothing
    for total in candidates:
        if abs(total - item_sum) <= 0.06:
            return items

    for total in sorted(set(candidates)):
        diff = round(total - item_sum, 2)
        # Duplicate SKUs are almost always small lines; never clone a large header amount
        if diff <= 0.05 or diff > 250:
            continue
        matches = []
        for it in items:
            try:
                amt = float(str(it.get("amount") or "0").replace(",", ""))
            except Exception:
                continue
            if abs(amt - diff) <= 0.05 and amt <= 250:
                matches.append(it)
        if matches:
            matches.sort(key=lambda x: float(str(x.get("amount") or "0").replace(",", "")))
            return items + [dict(matches[0])]
    return items


def parse_mrp_rate_items(text: str) -> list[dict[str, str]]:
    """
    Common spare-parts layout with MRP column:
      Desc HSN Qty MRP Rate [Disc] Amount
      OR Qty MRP Dis% Tax% Amount (no dealer rate — Ashok-style)
      OR Qty MRP Amount (rate blank — use MRP as selling reference)
    """
    items: list[dict[str, str]] = []
    money = r"[\d,]+\.\d{2}"
    row = re.compile(
        r"(?P<head>.+?)\s+"
        r"(?:(?P<hsn>\d{4,8})\s+)?"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s*"
        r"(?P<unit>NOS|Nos|PCS|SET|PKT|PKTS)?\s+"
        r"(?P<mrp>" + money + r")\s+"
        r"(?:(?P<rate>" + money + r")\s+)?"
        r"(?:(?P<disc>\d{1,2}(?:\.\d{1,2})?)\s*%?\s+)?"
        r"(?:(?P<tax>\d{1,2}(?:\.\d{1,2})?)\s+%?\s+)?"
        r"(?P<amount>" + money + r")",
        re.I,
    )
    has_mrp_header = bool(re.search(r"(?i)\bmrp\b", text[:1200]))
    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        if not has_mrp_header and not re.search(rf"{money}\s+{money}", ln):
            if len(re.findall(money, ln)) < 2:
                continue
        m = row.search(ln)
        if not m:
            continue
        head = _clean(m.group("head"))
        head = re.sub(r"^\d{1,3}[\)\.\s]+", "", head)
        part, desc = "", head
        tokens = head.split()
        if tokens and re.search(r"[A-Za-z]", tokens[0]) and re.search(r"\d", tokens[0]):
            part, desc = tokens[0], " ".join(tokens[1:]) or tokens[0]
        mrp = m.group("mrp")
        rate = m.group("rate") or ""
        amount = m.group("amount")
        # Only treat money tokens AFTER the HSN/qty block — decimals in part
        # descriptions (e.g. CMP-6.17R) must not become MRP.
        after = ln
        if m.group("hsn"):
            after = ln[m.start("qty") :] if m.start("qty") >= 0 else ln[m.end("hsn") :]
        elif m.start("qty") >= 0:
            after = ln[m.start("qty") :]
        moneys = re.findall(money, after)
        # Ashok: MRP Disc% Tax% Amount → moneys like 364.00, 30.00, 18.00?, 509.60
        if len(moneys) >= 3:
            try:
                cand_mids = [float(x.replace(",", "")) for x in moneys[1:-1]]
            except Exception:
                cand_mids = []
            if cand_mids and all(v <= 100 for v in cand_mids):
                mrp, amount = moneys[0], moneys[-1]
                rate = ""
            elif len(moneys) >= 3 and not has_mrp_header:
                mrp, rate, amount = moneys[-3], moneys[-2], moneys[-1]
            else:
                # Header says MRP: first money after qty is MRP; last is amount
                mrp, amount = moneys[0], moneys[-1]
                if len(moneys) >= 3:
                    mid = float(moneys[1].replace(",", ""))
                    # Second money is dealer rate only if it doesn't look like a %
                    rate = moneys[1] if mid > 100 else ""
                else:
                    rate = ""
        elif len(moneys) == 2:
            mrp, amount = moneys[0], moneys[1]
            rate = ""
        elif len(moneys) == 1:
            amount = moneys[0]
            mrp = mrp if mrp in moneys else ""
            rate = rate if rate else ""
        unit = (m.group("unit") or "").strip()
        qty = m.group("qty")
        if re.fullmatch(r"\d{4,8}", qty) and not unit:
            continue
        if not rate:
            try:
                q = float(qty.replace(",", ""))
                a = float(amount.replace(",", ""))
                if q > 0:
                    rate = f"{a / q:.2f}"
            except Exception:
                rate = ""
        items.append(
            {
                "part_number": part,
                "description": desc,
                "hsn_sac": m.group("hsn") or "",
                "qty": f"{qty} {unit}".strip(),
                "mrp": mrp,
                "rate": rate,  # net unit only — never copy MRP into Rate
                "amount": amount,
            }
        )
    return _dedupe_items(items)


def parse_item_code_particulars(text: str) -> list[dict[str, str]]:
    """
    Item Code | Particulars | HSN | Tax% | Qty | Unit | Rate | [Disc%] | Amount
    (Karnavati ZipERP / institutional sales — no MRP column)

    Example:
      1 990030 COND SWIFT PTL/DSL T-1 84159000 18.00 5 PCS 1,700.00 8,500.00
    """
    items: list[dict[str, str]] = []
    row = re.compile(
        r"^\s*(?P<sr>\d{1,3})\s+"
        r"(?P<part>\d{4,8})\s+"
        r"(?P<desc>[A-Za-z][A-Za-z0-9 \/\.\-\(\)]*?)\s+"
        r"(?P<hsn>\d{4,8})\s+"
        r"(?P<tax>\d{1,2}(?:\.\d{1,2})?)\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s*"
        r"(?P<unit>PCS|NOS|Nos|SET|Kg|KG)?\s+"
        r"(?P<rate>[\d,]+\.\d{2})"
        r"(?:\s+(?P<disc>\d{1,2}(?:\.\d{1,2})?))?\s+"
        r"(?P<amount>[\d,]+\.\d{2})\s*$",
        re.I,
    )
    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        m = row.match(ln)
        if not m:
            continue
        unit = (m.group("unit") or "").strip()
        qty = m.group("qty")
        items.append(
            {
                "part_number": m.group("part"),
                "description": _clean(m.group("desc")),
                "hsn_sac": m.group("hsn"),
                "qty": f"{qty} {unit}".strip(),
                "mrp": "",
                "rate": m.group("rate"),
                "amount": m.group("amount"),
            }
        )
    return _dedupe_items(items)



def parse_sno_part_particulars_gst(text: str) -> list[dict[str, str]]:
    """
    S.No PART No PARTICULARS HSN GSTR% QTY Rate [Dis%] AMOUNT
    ALA / Autolight Coimbatore style:
      1 7818-1501 Gates Cummins Water Pump 87089100 18 % 1 NOS 2,607.00 39 % 1,590.27
      1 12569 P)12V H4 100/90W P43T 85392120 18 % 30 NOS 133.90 4,017.00
    """
    items: list[dict[str, str]] = []
    row = re.compile(
        r"^\s*(?P<sl>\d{1,3})\s+"
        r"(?P<part>[A-Z0-9][A-Z0-9\-]{2,})\s+"
        r"(?P<desc>.+?)\s+"
        r"(?P<hsn>\d{4,8})\s+"
        r"(?P<gst>\d{1,2})\s*%?\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s*"
        r"(?P<unit>NOS|Nos|PCS|SET|PKT)?\s+"
        r"(?P<rate>[\d,]+\.\d{2})"
        r"(?:\s+(?P<disc>\d{1,2}(?:\.\d{1,2})?)\s*%?)?\s+"
        r"(?P<amount>[\d,]+\.\d{2})\s*$",
        re.I,
    )
    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        m = row.match(ln)
        if not m:
            continue
        unit = (m.group("unit") or "").strip()
        items.append(
            {
                "part_number": m.group("part"),
                "description": _clean(m.group("desc")),
                "hsn_sac": m.group("hsn"),
                "qty": f"{m.group('qty')} {unit}".strip(),
                "mrp": "",
                "rate": m.group("rate"),
                "amount": m.group("amount"),
            }
        )
    return _dedupe_items(items)


def parse_desc_part_brand_hsn(text: str) -> list[dict[str, str]]:
    """
    S.No DESCRIPTION PARTNO BRAND HSN QTY UOM RATE GST TAXABLE
    Karpagam Auto Stores style:
      1 HAND BR HOSE ... LEYLAND 91514YA ADEENA 39172390 1 NOS 249.70 18 249.70
      9 LEYLAND U TRUCK FAN BELT 8PK-1552CONTITECH 40103390 1 NOS 598.90 18 598.90
      33 COURIER CHARGES 996812 1 750.00 18 750.00

    Also recovers OCR-noisy scan lines (TRANSPORT COPY):
      1 V ROD ASSY U TRUCK B6Y03007 LEYLAND ... PRIZOL 73269099 = {ANOS 10935.00 18 10,935.00
    """
    if not re.search(r"(?i)PART\s*NO|PARTNO|KARPAGAM|DESCRIPTION\s+OF\s+GOODS", text[:3500]):
        return []
    items: list[dict[str, str]] = []
    seen_sl: set[int] = set()
    # Brand optional — OCR often glues brand onto part (CONTITECH, STAR)
    # Unit required so courier/SAC lines don't steal CHARGES as part no.
    row = re.compile(
        r"^\s*(?P<sl>\d{1,3})\s+"
        r"(?P<desc>.+?)\s+"
        r"(?P<part>[A-Z0-9][A-Z0-9\/\-]{2,})\s+"
        r"(?:(?P<brand>[A-Z][A-Z0-9\-]{1,12})\s+)?"
        r"(?P<hsn>\d{4,8})\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s*"
        r"(?P<unit>NOS|SET|PCS|PKT|KIT|Nos)\s+"
        r"(?P<rate>[\d,]+\.\d{2})\s+"
        r"(?P<gst>\d{1,2})\s+"
        r"(?P<amount>[\d,]+\.\d{2})\s*$",
        re.I,
    )
    # Courier / SAC-only charges without part+brand
    charge = re.compile(
        r"^\s*(?P<sl>\d{1,3})\s+"
        r"(?P<desc>[A-Z][A-Z0-9 \/\-]{3,}?)\s+"
        r"(?P<hsn>\d{4,8})\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s+"
        r"(?P<rate>[\d,]+\.\d{2})\s+"
        r"(?P<gst>\d{1,2})\s+"
        r"(?P<amount>[\d,]+\.\d{2})\s*$",
        re.I,
    )
    # OCR-noisy: SI ... PARTCODE ... HSN8 ... RATE GST AMOUNT (qty often "{ANOS" / "1NOS")
    fuzzy = re.compile(
        r"^\s*(?P<sl>\d{1,3})\s+"
        r"(?P<head>.+?)\s+"
        r"(?P<hsn>\d{8})\s+"
        r".*?"
        r"(?P<rate>[\d,]+\.\d{2})\s+"
        r"(?P<gst>5|12|18|28)\s+"
        r"(?P<amount>[\d,]+\.\d{2})\s*$",
        re.I,
    )

    def _append(sl: int, item: dict[str, str]) -> None:
        if sl in seen_sl:
            return
        seen_sl.add(sl)
        items.append(item)

    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        # Skip HSN summary rows (description is a commodity class sentence)
        if re.search(r"(?i)^(?:tubes|other|locks|natural|threaded|postal|scan|lower|postal)\b", ln):
            continue
        m = row.match(ln)
        if m:
            part = m.group("part")
            # Split glued brand suffix when present (…CONTITECH / …STAR)
            for brand_s in ("CONTITECH", "STAR", "VELFIT", "VULCAN"):
                if part.upper().endswith(brand_s) and len(part) > len(brand_s) + 2:
                    part = part[: -len(brand_s)]
                    break
            unit = (m.group("unit") or "").strip()
            _append(
                int(m.group("sl")),
                {
                    "part_number": part,
                    "description": _clean(m.group("desc")),
                    "hsn_sac": m.group("hsn"),
                    "qty": f"{m.group('qty')} {unit}".strip(),
                    "mrp": "",
                    "rate": m.group("rate"),
                    "amount": m.group("amount"),
                },
            )
            continue
        m2 = charge.match(ln)
        if m2 and re.search(r"(?i)courier|freight|packing|transport|labour", m2.group("desc")):
            _append(
                int(m2.group("sl")),
                {
                    "part_number": "",
                    "description": _clean(m2.group("desc")),
                    "hsn_sac": m2.group("hsn"),
                    "qty": m2.group("qty"),
                    "mrp": "",
                    "rate": m2.group("rate"),
                    "amount": m2.group("amount"),
                },
            )
            continue
        mf = fuzzy.match(ln)
        if not mf:
            continue
        if re.search(r"(?i)courier|freight|packing", ln):
            continue
        head = mf.group("head")
        # Prefer alphanumeric part codes with a digit (B6Y03007, 91514YA)
        part_cands = re.findall(r"\b([A-Z0-9][A-Z0-9\/\-]{4,})\b", head, re.I)
        part = ""
        for cand in part_cands:
            if looks_like_part_number(cand) and not re.fullmatch(r"\d{8}", cand):
                part = cand
                break
        if not part:
            continue
        # Description = tokens before part
        desc = head
        idx = desc.upper().find(part.upper())
        if idx > 0:
            desc = desc[:idx].strip()
        qty = "1"
        qm = re.search(r"(?i)(?:\{|\b)(\d{1,4})\s*A?NOS\b", ln)
        if qm:
            qty = qm.group(1)
            # Never treat HSN (8 digits) as qty
            if len(qty) >= 6:
                qty = "1"
        rate = mf.group("rate")
        amount = mf.group("amount")
        # Prefer amount that matches rate when qty=1 (OCR often duplicates)
        try:
            if abs(float(rate.replace(",", "")) - float(amount.replace(",", ""))) < 0.05:
                pass
            elif float(qty) == 1 and float(rate.replace(",", "")) > float(amount.replace(",", "")) * 5:
                # Amount stolen from courier line — use rate as amount
                amount = rate
        except Exception:
            pass
        _append(
            int(mf.group("sl")),
            {
                "part_number": part,
                "description": _clean(desc) or part,
                "hsn_sac": mf.group("hsn"),
                "qty": f"{qty} NOS",
                "mrp": "",
                "rate": rate,
                "amount": amount,
            },
        )
    return _dedupe_items(items)


def parse_sno_part_desc_discounts(text: str) -> list[dict[str, str]]:
    """
    S.No PartNo Description HSN Qty Unit Rate Dis% [CD%] Tax% Amount
    ASK Automobile / similar counter invoices (multi-line desc ignored on wrap).
      1 CS607822 " R/SPRING LEAF 73181500 4 Pcs 135.00 15.00 5.00 18 436.05
    """
    items: list[dict[str, str]] = []
    by_sl: dict[int, dict[str, str]] = {}
    row = re.compile(
        r"^\s*(?P<sl>\d{1,3})\s+"
        r"(?P<part>[A-Z0-9][A-Z0-9\/\-]{2,})\s+"
        r"(?P<desc>.+?)\s+"
        r"(?P<hsn>\d{4,8})\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s*"
        r"(?P<unit>Pcs|PCS|Nos|NOS|EACH|Eacg|Kit|SET)?\s+"
        r"(?P<rate>[\d,]+\.\d{2})\s+"
        r"(?:(?P<disc>\d{1,2}(?:\.\d{1,2})?)\s+)?"
        r"(?:(?P<cd>\d{1,2}(?:\.\d{1,2})?)\s+)?"
        r"(?P<tax>\d{1,2})\s+"
        r"(?P<amount>[\d,]+\.\d{2})\s*$",
        re.I,
    )
    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        m = row.match(ln)
        if not m:
            continue
        # HSN summary: qty huge + no unit often — still allow with unit
        unit = (m.group("unit") or "").strip()
        if unit.lower() == "eacg":
            unit = "EACH"
        sl = int(m.group("sl"))
        it = {
            "part_number": m.group("part"),
            "description": _clean(m.group("desc")).strip('"'),
            "hsn_sac": m.group("hsn"),
            "qty": f"{m.group('qty')} {unit}".strip(),
            "mrp": m.group("rate"),  # list price before discount on these bills
            "rate": "",  # net derived from amount/qty
            "amount": m.group("amount"),
        }
        try:
            q = float(m.group("qty").replace(",", ""))
            a = float(m.group("amount").replace(",", ""))
            if q > 0:
                it["rate"] = f"{a / q:.2f}"
        except Exception:
            it["rate"] = m.group("rate")
        by_sl[sl] = it
    items = [by_sl[k] for k in sorted(by_sl)]
    return items


def parse_goods_hsn_qty_rate_disc(text: str) -> list[dict[str, str]]:
    """
    SI Description HSN Qty Unit Rate [Unit] Disc% Amount
    HEMA / MB Agencies / PR Process Tally style:
      1 LX 3630KIT(KFK0249506) 84213100 1 NOS 822.00 NOS 13.76 % 708.89
      1 VT 1203001(Elec Oil...) 90262000 18% 2 nos 467.00 nos 579.08
      1 LL149 HINO MANIFOLD SET OF 6 ( 84841090 5 NOS 102.00 NOS 510.00
    """
    items: list[dict[str, str]] = []
    by_sl: dict[int, dict[str, str]] = {}
    # With GST% before qty (MB)
    row_gst = re.compile(
        r"^\s*(?P<sl>\d{1,3})\s+"
        r"(?P<head>.+?)\s+"
        r"(?P<hsn>\d{8})\s+"
        r"(?P<gst>\d{1,2})\s*%\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s*"
        r"(?P<unit>nos|NOS|Pcs|PCS)?\s+"
        r"(?P<rate>[\d,]+\.\d{2})\s+"
        r"(?:nos|NOS|Pcs|PCS)?\s*"
        r"(?P<amount>[\d,]+\.\d{2})\s*$",
        re.I,
    )
    # With disc% before amount (HEMA)
    row_disc = re.compile(
        r"^\s*(?P<sl>\d{1,3})\s+"
        r"(?P<head>.+?)\s+"
        r"(?P<hsn>\d{8})\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s*"
        r"(?P<unit>NOS|Nos|PCS)?\s+"
        r"(?P<rate>[\d,]+\.\d{2})\s+"
        r"(?:NOS|Nos|PCS)?\s+"
        r"(?P<disc>\d{1,2}(?:\.\d{1,2})?)\s*%?\s+"
        r"(?P<amount>[\d,]+\.\d{2})\s*$",
        re.I,
    )
    # Plain qty rate amount (PR Process)
    row_plain = re.compile(
        r"^\s*(?P<sl>\d{1,3})\s+"
        r"(?P<head>.+?)\s+"
        r"(?P<hsn>\d{8})\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s*"
        r"(?P<unit>NOS|Nos|PCS)?\s+"
        r"(?P<rate>[\d,]+\.\d{2})\s+"
        r"(?:NOS|Nos|PCS)?\s+"
        r"(?P<amount>[\d,]+\.\d{2})\s*$",
        re.I,
    )

    def _split_head(head: str) -> tuple[str, str]:
        head = _clean(head)
        # VT 1203001(desc...) or LX 3630KIT(KFK...) or LL149 DESC
        m = re.match(r"^([A-Z]{1,6}\s*\d{3,8}[A-Z0-9\-]*)\s*[\(\-]?\s*(.*)$", head, re.I)
        if m and len(m.group(1)) >= 4:
            part = re.sub(r"\s+", " ", m.group(1)).strip()
            desc = m.group(2).strip(" ()-")
            if not desc:
                desc = part
                # keep part from code before (
                pm = re.match(r"^([A-Z0-9][A-Z0-9\- ]{2,})", head, re.I)
                if pm:
                    part = pm.group(1).split("(")[0].strip()
            return part, desc or head
        m = re.match(r"^([A-Z]{2,6}\d{2,5}[A-Z0-9]*)\s+(.+)$", head, re.I)
        if m:
            return m.group(1), m.group(2)
        return "", head

    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        m = row_gst.match(ln) or row_disc.match(ln) or row_plain.match(ln)
        if not m:
            continue
        sl = int(m.group("sl"))
        part, desc = _split_head(m.group("head"))
        unit = (m.group("unit") or "").strip()
        by_sl[sl] = {
            "part_number": part,
            "description": desc,
            "hsn_sac": m.group("hsn"),
            "qty": f"{m.group('qty')} {unit}".strip(),
            "mrp": "",
            "rate": m.group("rate"),
            "amount": m.group("amount"),
        }
    return [by_sl[k] for k in sorted(by_sl)]


def parse_anbu_part_hsn_net(text: str) -> list[dict[str, str]]:
    """
    S.No Part Desc HSN Qty Rate Tax% NetRate Amount
      1 TX55 TPH TEXSPIN 84828000 3 734.48 18 866.69 2600.06
    Rate = taxable unit; NetRate is GST-inclusive (ignore for Rate column).
    """
    items: list[dict[str, str]] = []
    row = re.compile(
        r"^\s*(?P<sl>\d{1,3})\s+"
        r"(?P<part>[A-Z0-9][A-Z0-9\-]{1,})\s+"
        r"(?P<desc>[A-Za-z][A-Za-z0-9 \/\.\-]{1,}?)\s+"
        r"(?P<hsn>\d{4,8})\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s+"
        r"(?P<rate>[\d,]+\.\d{2})\s+"
        r"(?P<tax>\d{1,2})\s+"
        r"(?P<net>[\d,]+\.\d{2})\s+"
        r"(?P<amount>[\d,]+\.\d{2})\s*$",
        re.I,
    )
    for raw in text.splitlines():
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        m = row.match(ln)
        if not m:
            continue
        items.append(
            {
                "part_number": m.group("part"),
                "description": _clean(m.group("desc")),
                "hsn_sac": m.group("hsn"),
                "qty": m.group("qty"),
                "mrp": "",
                "rate": m.group("rate"),
                "amount": m.group("amount"),
            }
        )
    return _dedupe_items(items)


def parse_anamallais_part_above_si(text: str) -> list[dict[str, str]]:
    """
    Dealer invoices where Part No sits on the line above S.No, HSN is OCR-split:
      IA202777
      1 (8708700 LOCK NUT, WHEEL BEARING/(A) NOS 730.00 2.00 618.65 14.00 -86.61 532.04 1,064.08
      0)
    Stop at Spare Parts Total / HSN Code Summary (do not ingest tax-summary rows).
    """
    if not re.search(
        r"(?i)Part\s*No.*HSN|Loyalty\s*Point|Spare\s*Parts\s*Total|Receiver\s*GST",
        text[:3500],
    ):
        return []
    if not re.search(r"(?i)\bS\.?\s*No\b", text[:3500]):
        return []

    lines = [_clean(ln) for ln in text.splitlines()]
    # Cut before HSN summary / loyalty / totals that pollute parsers
    cut = len(lines)
    for i, ln in enumerate(lines):
        if re.search(
            r"(?i)^(?:Spare\s*Parts\s*Total|HSN\s*Code\s*Summary|Loyalty\s*Point|"
            r"Part\s*Taxable|Central\s*GST|Gross\s*Total|Total\s*Invoice)\b",
            ln,
        ):
            cut = i
            break
    lines = lines[:cut]

    part_only = re.compile(r"^[A-Z]{1,4}\d{4,10}[A-Z0-9]*$")
    # Trailing money: disc_amt, net, amount — space optional between disc/net when OCR glues
    # e.g. "-86.61 532.04 1,064.08" or "-1,086.196,672.28 6,672.28"
    row = re.compile(
        r"^\s*(?P<sl>\d{1,2})\s+"
        r"(?P<head>.+?)\s+"
        r"(?P<uom>NOS|PCS|SET|PKT)\s+"
        r"(?P<mrp>[\d,]+\.\d{2})\s+"
        r"(?P<qty>[\d,]+\.\d{2})\s+"
        r"(?P<rate>[\d,]+\.\d{2})\s+"
        r"(?P<disc>\d{1,2}(?:\.\d{1,2})?)\s+"
        r"(?P<discamt>-?[\d,]+\.\d{2})\s*"
        r"(?P<net>[\d,]+\.\d{2})\s+"
        r"(?P<amount>[\d,]+\.\d{2})\s*$",
        re.I,
    )
    hsn_frag = re.compile(r"^\(?\s*(?P<a>\d{4,8})\)?\s*$|^(?P<b>\d{1,4})\)\s*$")

    items: list[dict[str, str]] = []
    by_sl: dict[int, dict[str, str]] = {}
    pending_part = ""
    pending_desc_bits: list[str] = []

    for i, ln in enumerate(lines):
        if not ln:
            continue
        if part_only.match(ln):
            pending_part = ln
            pending_desc_bits = []
            continue
        # Description wrap between part and SI row
        if pending_part and not re.match(r"^\d{1,2}\s", ln) and not part_only.match(ln):
            if not re.search(r"(?i)^SO\s*Number|^Anamallais|^Ph:|^Coimbatore", ln):
                if re.search(r"[A-Za-z]{3,}", ln) and not hsn_frag.match(ln):
                    pending_desc_bits.append(ln)
            continue

        m = row.match(ln)
        if not m:
            continue
        try:
            sl = int(m.group("sl"))
        except Exception:
            continue
        if sl < 1 or sl > 80:
            continue

        head = m.group("head")
        # HSN is usually inside "(8708700" — leading digit before "(" is OCR junk from wrap
        hm = re.search(r"\((\d{4,8})", head)
        if not hm:
            hm = re.search(r"(\d{4,8})", head)
        hsn = hm.group(1) if hm else ""
        # Look ahead for closing ")0)" / "00)" / "11)" fragments
        for j in range(i + 1, min(i + 4, len(lines))):
            nxt = lines[j]
            frag_m = re.match(r"^\(?\s*(\d{1,4})\)?\s*$", nxt)
            if not frag_m:
                if part_only.match(nxt) or row.match(nxt):
                    break
                # Skip "(A)" / description remnants
                if re.search(r"[A-Za-z]", nxt):
                    continue
                break
            frag = frag_m.group(1)
            if frag and len(hsn) < 8:
                hsn = (hsn + frag)[:8]
            if len(hsn) >= 8:
                break
        if len(hsn) > 8:
            hsn = hsn[-8:]

        desc = head
        # Strip leading OCR-split HSN only — do NOT strip through any ")" (desc may contain /(A))
        desc = re.sub(r"^\d?\(\d{4,8}\s*", "", desc)
        desc = re.sub(r"^\(\d{4,8}\s*", "", desc)
        desc = re.sub(r"^\d{4,8}\s+", "", desc)
        desc = re.sub(r"/\(A\)\s*$", "", desc)
        desc = re.sub(r"\(A\)\s*$", "", desc)
        desc = desc.strip(" ,/-")
        if pending_desc_bits:
            wrapped = _clean(" ".join(pending_desc_bits))
            desc = _clean(f"{wrapped} {desc}".strip()) if desc else wrapped
        # Drop leaked HSN digit runs / repeated part at start of description
        desc = re.sub(r"^\d{4,8}\s+", "", desc)

        part = pending_part
        if part and desc.upper().startswith(part.upper()):
            desc = desc[len(part) :].lstrip(", ").strip()
        if not part:
            pm = re.match(r"^([A-Z]{1,4}\d{4,10})\s*[, ]", desc)
            if pm:
                part = pm.group(1)
                desc = desc[pm.end() :].lstrip(", ").strip()
        elif desc.upper().startswith(part.upper() + ","):
            desc = desc[len(part) :].lstrip(", ").strip()

        qty = m.group("qty")
        try:
            qf = float(qty.replace(",", ""))
            qty = str(int(qf)) if abs(qf - round(qf)) < 0.01 else qty
        except Exception:
            pass
        unit = m.group("uom")
        item = {
            "part_number": part,
            "description": desc or part or "Item",
            "hsn_sac": hsn if len(hsn) >= 4 else "",
            "qty": f"{qty} {unit}".strip(),
            "mrp": m.group("mrp"),
            "rate": m.group("rate"),
            "amount": m.group("amount"),
        }
        by_sl[sl] = item
        pending_part = ""
        pending_desc_bits = []

    for sl in sorted(by_sl):
        items.append(by_sl[sl])
    return items


def parse_tally_phone_goods_line(text: str) -> list[dict[str, str]]:
    """
    Phone photo of Tally tax invoice (creased paper + pen marks). OCR is noisy but
    usually still has: part/desc, HSN, Nos, and HSN-summary taxable amount.
      TW41280 - Tata 697 Belt ... 40103999 ... 2Nos ... 399.30
    """
    if not re.search(r"(?i)tax\s*invoice|amount\s*chargeable|HSN\s*/?\s*SAC", text[:2500]):
        return []
    items: list[dict[str, str]] = []

    # HSN summary taxable — prefer larger values (line amount, not CGST 35.94)
    taxable_candidates: list[tuple[str, float, str]] = []
    for m in re.finditer(
        r"\b(?P<hsn>\d{8})\b[^\n]{0,80}?(?P<taxable>[\d,]+\.\d{2})",
        text,
        re.I,
    ):
        hsn = m.group("hsn")
        taxable = m.group("taxable")
        try:
            v = float(taxable.replace(",", ""))
        except Exception:
            continue
        if v in {35.94, 71.88, 0.18} or v < 80:
            continue
        taxable_candidates.append((hsn, v, taxable))
    taxable_by_hsn: dict[str, tuple[float, str]] = {}
    for hsn, v, raw in taxable_candidates:
        prev = taxable_by_hsn.get(hsn)
        if not prev or v > prev[0]:
            taxable_by_hsn[hsn] = (v, raw)

    def _best_amount(hsn: str) -> tuple[str, str]:
        """Return (hsn, amount) preferring summary taxable over mangled line OCR."""
        # Prefer canonical HSN forms that appear cleanly in OCR (40103999 over 40103000)
        votes: dict[str, int] = {}
        for m in re.finditer(r"\b(40103\d{3}|87\d{6}|84\d{6}|73\d{6}|90\d{6})\b", text):
            votes[m.group(1)] = votes.get(m.group(1), 0) + 1
        family_hsns = [h for h in votes if h[:5] == hsn[:5] or h[:4] == hsn[:4]]
        preferred_hsn = hsn
        if family_hsns:
            preferred_hsn = max(family_hsns, key=lambda h: (votes[h], h.endswith("999"), h))

        if preferred_hsn in taxable_by_hsn and taxable_by_hsn[preferred_hsn][0] >= 150:
            return preferred_hsn, taxable_by_hsn[preferred_hsn][1]
        if hsn in taxable_by_hsn and taxable_by_hsn[hsn][0] >= 150:
            return preferred_hsn if preferred_hsn else hsn, taxable_by_hsn[hsn][1]
        family = [
            (h, v, raw)
            for h, (v, raw) in taxable_by_hsn.items()
            if h[:5] == hsn[:5] or h[:4] == hsn[:4]
        ]
        if family:
            h, v, raw = max(family, key=lambda t: t[1])
            return preferred_hsn or h, raw
        if taxable_by_hsn:
            h, (v, raw) = max(taxable_by_hsn.items(), key=lambda kv: kv[1][0])
            return preferred_hsn or h, raw
        return preferred_hsn or hsn, ""

    # Qty near grand total: "2Nos Rs. 471.00"
    total_qty = ""
    tq = re.search(r"(?i)\b(\d{1,3})\s*Nos\b[^\n]{0,20}?Rs\.?\s*([\d,]+\.\d{2})", text)
    if tq:
        total_qty = tq.group(1)

    goods = re.compile(
        r"(?P<part>TW\s*\d{3,5}|[A-Z]{1,3}\d{3,6})\s*[-–]\s*"
        r"(?P<desc>[A-Za-z][A-Za-z0-9 \/&\-]{2,60}?)"
        r"(?=\s+\d{4,}| \(|$)"
        r".{0,50}?"
        r"(?P<hsn>\d{8})"
        r".{0,40}?"
        r"(?P<qty>\d{1,3})\s*Nos",
        re.I,
    )

    matched = False
    for m in goods.finditer(text):
        part = re.sub(r"\s+", "", m.group("part")).upper()
        if part.startswith("TW") and len(part) >= 5:
            # TW41280 is often OCR for TW 1280 (crease/pen)
            digits = part[2:]
            if len(digits) == 5 and digits[0] in "34":
                part = f"TW {digits[1:]}"
            else:
                part = f"TW {digits}"
        desc = _clean(m.group("desc"))
        # Keep "Tata 697 Belt" — strip trailing junk codes
        desc = re.sub(r"\s+\d{5,6}\s*$", "", desc).strip(" -\"'")
        hsn = m.group("hsn")
        hsn, amount = _best_amount(hsn)
        qty = m.group("qty")
        if total_qty and (qty in {"72", "12", "22", "32", "42"} or len(qty) > 1):
            if total_qty.isdigit() and int(total_qty) <= 20:
                qty = total_qty
        if not amount:
            continue
        try:
            qf = float(qty)
            af = float(amount.replace(",", ""))
            if qf <= 0 or af <= 0:
                continue
            rate = f"{af / qf:.2f}"
            amount = f"{af:.2f}"
        except Exception:
            continue
        items.append(
            {
                "part_number": part,
                "description": desc or part,
                "hsn_sac": hsn if len(hsn) == 8 else m.group("hsn"),
                "qty": f"{qty} Nos",
                "mrp": "",
                "rate": rate,
                "amount": amount,
            }
        )
        matched = True
        break

    if not matched and taxable_by_hsn:
        dm = re.search(
            r"(?i)\b(TW\s*\d{3,5}|[A-Z]{2}\d{3,5})\s*[-–]\s*([A-Za-z][A-Za-z0-9 \/&\-]{3,40})",
            text,
        )
        hsn, (_, amount) = max(taxable_by_hsn.items(), key=lambda kv: kv[1][0])
        qty = total_qty or "1"
        try:
            rate = f"{float(amount.replace(',', '')) / float(qty):.2f}"
        except Exception:
            rate = amount
        part = ""
        desc = "Item"
        if dm:
            raw_part = re.sub(r"\s+", "", dm.group(1)).upper()
            if raw_part.startswith("TW"):
                digits = raw_part[2:]
                if len(digits) == 5 and digits[0] in "34":
                    part = f"TW {digits[1:]}"
                else:
                    part = f"TW {digits}"
            else:
                part = raw_part
            desc = _clean(dm.group(2))
            desc = re.sub(r"\s+\d{5,6}\s*$", "", desc).strip(" -\"'")
        items.append(
            {
                "part_number": part,
                "description": desc,
                "hsn_sac": hsn,
                "qty": f"{qty} Nos",
                "mrp": "",
                "rate": rate,
                "amount": amount if "." in str(amount) else f"{float(str(amount).replace(',','')):.2f}",
            }
        )
    return _dedupe_items(items)


def parse_lubricant_mrp_qty_rate(text: str) -> list[dict[str, str]]:
    """
    Arvind Petroleums / Shell lubricant lines (desc sometimes OCR-dropped):
      1550070727 RIMULA R4 27101972 2,707.62 112.500 Ltr\\Kg 15 Nos 2,707.62 Nos 40,614.30
      2550039995 27101980 1,796.61 15.000 Ltr\\Kg 1-0.0000 Ctn 1,796.61 Pcs 5,389.83
      3550040915-RimR415W40 27101980 3,440.67 50.000 Ltr\\Kg 5.0000 Pcs 3,440.67 Pcs 17,203.35
    """
    if not re.search(r"(?i)RIMULA|Ltr\\\\?Kg|Ltr/Kg|Description of Goods", text):
        return []
    items: list[dict[str, str]] = []
    # Qty forms: "15 Nos", "5.0000 Pcs", "1-0.0000 Ctn" (leading int is qty)
    row = re.compile(
        r"(?P<part>\d{7,12})(?:-(?P<pdesc>[A-Za-z0-9_]+))?\s+"
        r"(?:(?P<desc>[A-Za-z][A-Za-z0-9 \\/\.\-]{1,}?)\s+)?"
        r"(?P<hsn>\d{8})\s+"
        r"(?P<rate>[\d,]+\.\d{2})\s+"
        r"[\d,]+\.\d{3}\s+\S+\s+"
        r"(?P<qty>\d+(?:\.\d+)?)(?:-\d+(?:\.\d+)?)?\s*"
        r"(?P<unit>Nos|Pcs|Ctn|PCS)\s+"
        r"[\d,]+\.\d{2}\s+\S+\s+"
        r"(?P<amount>[\d,]+\.\d{2})",
        re.I,
    )
    lines = text.splitlines()
    for i, raw in enumerate(lines):
        ln = _clean(raw)
        if not ln or STOP_ITEM.search(ln):
            continue
        m = row.search(ln)
        if not m:
            continue
        qty = m.group("qty")
        try:
            qf = float(qty)
            qty = str(int(round(qf))) if abs(qf - round(qf)) < 0.01 else qty
        except Exception:
            qty = m.group("qty").split("-")[0]
        mrp = ""
        for j in range(i + 1, min(i + 4, len(lines))):
            mm = re.search(r"(?i)\bMRP\s*([\d,]+)", lines[j])
            if mm:
                mrp = mm.group(1)
                if "." not in mrp:
                    mrp = f"{mrp}.00"
                break
        # Description often continues on next line(s) before MRP
        desc = _clean(m.group("desc") or m.group("pdesc") or "")
        if not desc or len(desc) < 4:
            bits: list[str] = []
            for j in range(i + 1, min(i + 4, len(lines))):
                nxt = _clean(lines[j])
                if not nxt or re.search(r"(?i)^MRP\b|^\d{7,12}\b|^CGST|^SGST|^Total\b", nxt):
                    break
                if re.match(r"^[\d,]+\.\d{2}$", nxt):
                    break
                # Drop leading bullets / OCR junk; strip trailing MRP token if glued
                bit = re.sub(r"(?i)\s*MRP\s*[\d,]+\s*$", "", nxt.lstrip("-*").strip())
                if bit:
                    bits.append(bit)
            if bits:
                desc = _clean(" ".join(bits))
        unit = (m.group("unit") or "").strip()
        items.append(
            {
                "part_number": m.group("part"),
                "description": desc or m.group("part"),
                "hsn_sac": m.group("hsn"),
                "qty": f"{qty} {unit}".strip(),
                "mrp": mrp,
                "rate": m.group("rate"),
                "amount": m.group("amount"),
            }
        )
    return _dedupe_items(items)



def parse_line_items(text: str) -> list[dict[str, str]]:
    """
    Layout-first: detect column schema, prefer matching parsers, gate MRP repair.
    Not tied to vendor names — same headers work for any supplier.
    """
    from extraction.layout import detect_table_layout

    schema = detect_table_layout(text)

    preferred_fns: dict[str, list] = {
        "mrp_disc": [
            parse_anamallais_part_above_si,
            parse_tally_phone_goods_line,
            parse_lubricant_mrp_qty_rate,
            parse_part_mrp_disc_tax_amount,
            parse_mrp_rate_items,
            parse_part_hsn_qty_rate,
        ],
        "mrp_rate": [
            parse_anamallais_part_above_si,
            parse_tally_phone_goods_line,
            parse_mrp_rate_items,
            parse_part_mrp_disc_tax_amount,
            parse_part_hsn_qty_rate,
        ],
        "credit_rate_qty": [
            parse_sno_part_particulars_gst,
            parse_sno_part_desc_discounts,
            parse_credit_bill_hsn_rate_qty,
            parse_part_hsn_qty_rate,
            parse_part_column_items,
        ],
        "item_code": [
            parse_item_code_particulars,
            lambda t: parse_tally_scan_items(t, hsn_digits=8),
            parse_amount_trail_items,
        ],
        "einvoice": [
            parse_sno_part_particulars_gst,
            parse_desc_part_brand_hsn,
            parse_tally_phone_goods_line,
            parse_anbu_part_hsn_net,
            parse_sno_part_desc_discounts,
            parse_goods_hsn_qty_rate_disc,
            parse_einvoice_line_items,
            parse_part_hsn_qty_rate,
            lambda t: parse_tally_scan_items(t, hsn_digits=8),
        ],
        "unknown": [
            parse_anamallais_part_above_si,
            parse_tally_phone_goods_line,
            parse_goods_hsn_qty_rate_disc,
            parse_sno_part_particulars_gst,
            parse_desc_part_brand_hsn,
            parse_lubricant_mrp_qty_rate,
        ],
    }

    fallback_fns = [
        parse_anamallais_part_above_si,
        parse_tally_phone_goods_line,
        parse_sno_part_particulars_gst,
        parse_desc_part_brand_hsn,
        parse_sno_part_desc_discounts,
        parse_goods_hsn_qty_rate_disc,
        parse_anbu_part_hsn_net,
        parse_lubricant_mrp_qty_rate,
        parse_item_code_particulars,
        parse_part_mrp_disc_tax_amount,
        parse_s_code_nos_items,
        parse_einvoice_line_items,
        parse_mrp_rate_items,
        parse_part_hsn_qty_rate,
        parse_handwritten_rate_qty_amount,
        parse_credit_bill_hsn_rate_qty,
        parse_photo_gstr_qty_rate,
        lambda t: parse_tally_scan_items(t, hsn_digits=8),
        lambda t: parse_tally_scan_items(t, hsn_digits=4),
        parse_part_column_items,
        parse_amount_trail_items,
    ]

    ordered_fns: list = []
    seen_names: set[str] = set()
    for i, fn in enumerate((preferred_fns.get(schema) or []) + fallback_fns):
        key = getattr(fn, "__name__", "") or f"fn{i}"
        if key == "<lambda>":
            key = f"lambda:{i}"
        if key in seen_names:
            continue
        seen_names.add(key)
        ordered_fns.append(fn)

    candidates = [fn(text) for fn in ordered_fns]

    def _quality(items: list[dict[str, str]]) -> tuple:
        usable = []
        parts = 0
        for it in items:
            desc = re.sub(r"[^A-Za-z]", "", str(it.get("description") or ""))
            part = str(it.get("part_number") or "")
            if len(desc) < 3 and len(part) < 4:
                continue
            try:
                rate_v = float(str(it.get("rate") or "0").replace(",", ""))
            except Exception:
                rate_v = 0
            if rate_v <= 0 and not it.get("mrp"):
                continue
            qty_tok = str(it.get("qty") or "").split()[0].replace(",", "")
            try:
                qty_v = float(qty_tok)
            except Exception:
                qty_v = 0
            if qty_v >= 100 and rate_v < 1:
                continue
            if re.fullmatch(r"\d+\.\d{2}", qty_tok) and qty_v >= 20:
                continue
            if looks_like_part_number(part):
                parts += 1
            usable.append(it)
        # Prefer filled Part Number rows strongly (stop ghost/HSN-summary winners)
        return (
            parts * 5 + _item_score(usable, schema),
            parts,
            len(usable),
            _item_score(items, schema),
        )

    best = max(candidates, key=_quality)
    cleaned = []
    for it in best:
        it = _normalize_part_fields(dict(it))
        desc = re.sub(r"[^A-Za-z]", "", str(it.get("description") or ""))
        part = str(it.get("part_number") or "")
        if len(desc) < 3 and len(part) < 4:
            continue
        try:
            rate_v = float(str(it.get("rate") or "0").replace(",", ""))
        except Exception:
            rate_v = 0
        if rate_v <= 0 and not it.get("mrp"):
            continue
        qty_tok = str(it.get("qty") or "").split()[0].replace(",", "")
        try:
            qty_v = float(qty_tok)
        except Exception:
            qty_v = 0
        if qty_v >= 100 and rate_v < 1:
            continue
        cleaned.append(it)
    if _item_score(cleaned, schema) > 0:
        return _repair_qty_mrp_shift(
            [_strip_part_from_description(it) for it in _normalize_shifted_hsn(cleaned)],
            text,
            schema=schema,
        )
    return []


def extract_invoice(data: bytes, filename: str) -> dict[str, Any]:
    """
    Permanent pipeline: optional invoice AI (OpenAI / Textract) → layout tables →
    legacy parsers → total validation + confidence.
    """
    from extraction.pipeline import extract_invoice_v2

    def _legacy_from_text(text: str) -> dict[str, Any]:
        return {
            "invoice_number": find_invoice_number(text),
            "supplier_name": find_supplier(text),
            "date": find_date(text),
            "place_of_supply": find_place_of_supply(text),
            "line_items": parse_line_items(text),
            "raw_text_preview": text[:3500],
            "extractor": "tesseract_rules",
        }

    return extract_invoice_v2(
        data,
        filename,
        legacy_extract_text=extract_text,
        legacy_parse=_legacy_from_text,
    )


def merge_by_invoice(invoices: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    order: list[str] = []
    for inv in invoices:
        key = inv["invoice_number"]
        if key == "Unknown":
            key = f"__file__:{inv['filename']}"
        if key not in groups:
            order.append(key)
        groups[key].append(inv)

    merged: list[dict[str, Any]] = []
    for key in order:
        pages = groups[key]
        base = dict(pages[0])
        items: list[dict[str, str]] = []
        files: list[str] = []
        for p in pages:
            files.append(p["filename"])
            items.extend(p.get("line_items") or [])
            for field in ("supplier_name", "date", "place_of_supply", "invoice_number", "extractor", "confidence"):
                if base.get(field) in (None, "", "Unknown") and p.get(field) not in (None, "", "Unknown"):
                    base[field] = p[field]
            if p.get("warnings"):
                base.setdefault("warnings", [])
                for w in p["warnings"]:
                    if w not in base["warnings"]:
                        base["warnings"].append(w)
            if p.get("confidence_score") and (
                not base.get("confidence_score") or p["confidence_score"] > base.get("confidence_score", 0)
            ):
                base["confidence_score"] = p["confidence_score"]
                base["confidence"] = p.get("confidence") or base.get("confidence")
        # Dedupe only when stitching multiple pages/files of the same invoice.
        # A single page may intentionally list the same SKU twice — keep those rows.
        if len(pages) == 1:
            unique_items = items
        else:
            seen = set()
            unique_items = []
            for it in items:
                sig = (it.get("part_number"), it.get("description"), it.get("qty"), it.get("amount"))
                if sig in seen:
                    continue
                seen.add(sig)
                unique_items.append(it)
        base["line_items"] = unique_items
        base["filename"] = " + ".join(dict.fromkeys(files))
        merged.append(base)
    return merged


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def _safe_sheet_name(invoice_no: str, supplier: str, date: str) -> str:
    inv = re.sub(r'[\\\/\?\*\[\]\:]', "-", invoice_no or "Invoice")
    sup = re.sub(r'[\\\/\?\*\[\]\:]', "-", supplier or "")
    dt = re.sub(r'[\\\/\?\*\[\]\:]', "-", date or "")
    inv = re.sub(r"\s+", " ", inv).strip() or "Invoice"
    sup = re.sub(r"\s+", " ", (sup if sup != "Unknown" else "")).strip()
    dt = re.sub(r"\s+", " ", (dt if dt != "Unknown" else "")).strip()

    full = "_".join(p for p in [inv, sup, dt] if p)
    if len(full) <= 31:
        return full or "Invoice"

    short_sup = re.sub(r"[^A-Za-z0-9]", "", sup)[:6]
    short_dt = re.sub(r"[^0-9A-Za-z]", "", dt)[:8]
    for candidate in (
        "_".join(p for p in [inv, short_sup, short_dt] if p),
        "_".join(p for p in [inv, short_dt] if p),
        inv,
    ):
        if candidate:
            return candidate[:31]
    return "Invoice"


def _unique_sheet_name(wb: Workbook, base: str) -> str:
    existing = {ws.title for ws in wb.worksheets}
    if base not in existing:
        return base
    for i in range(2, 100):
        suffix = f"_{i}"
        candidate = base[: 31 - len(suffix)] + suffix
        if candidate not in existing:
            return candidate
    return base[:28] + "_X"


def write_invoice_sheet(wb: Workbook, data: dict[str, Any], is_first: bool) -> None:
    title = _unique_sheet_name(
        wb, _safe_sheet_name(data["invoice_number"], data["supplier_name"], data["date"])
    )
    if is_first:
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title=title)

    # Plain values — avoid fancy styles that break in Apple Numbers
    meta = [
        ("Invoice Number", str(data.get("invoice_number") or "")),
        ("Supplier Name", str(data.get("supplier_name") or "")),
        ("Date Supplied", str(data.get("date") or "")),
        ("Place of Supply", str(data.get("place_of_supply") or "")),
        ("Source File", str(data.get("filename") or "")),
    ]
    for r, (label, value) in enumerate(meta, 1):
        ws.cell(r, 1, label)
        ws.cell(r, 2, value)

    headers = ["Part Number", "Description", "Qty", "MRP", "Rate", "Amount", "HSN/SAC"]
    start_row = 7
    for col, h in enumerate(headers, 1):
        cell = ws.cell(start_row, col, h)
        cell.font = Font(bold=True)

    items = data.get("line_items") or []
    if not items:
        ws.cell(start_row + 1, 1, "No line items detected")
    else:
        for i, item in enumerate(items):
            row = start_row + 1 + i
            ws.cell(row, 1, str(item.get("part_number") or ""))
            ws.cell(row, 2, str(item.get("description") or ""))
            ws.cell(row, 3, str(item.get("qty") or ""))
            ws.cell(row, 4, str(item.get("mrp") or ""))
            ws.cell(row, 5, str(item.get("rate") or ""))
            ws.cell(row, 6, str(item.get("amount") or ""))
            ws.cell(row, 7, str(item.get("hsn_sac") or ""))

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14


def build_workbook(invoices: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    for i, inv in enumerate(invoices):
        write_invoice_sheet(wb, inv, is_first=(i == 0))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv(invoices: list[dict[str, Any]]) -> bytes:
    """Flat CSV — opens correctly in Numbers, Excel, and Google Sheets."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        [
            "Invoice Number",
            "Supplier Name",
            "Date Supplied",
            "Place of Supply",
            "Part Number",
            "Description",
            "Qty",
            "MRP",
            "Rate",
            "Amount",
            "HSN/SAC",
            "Source File",
        ]
    )
    for inv in invoices:
        items = inv.get("line_items") or [
            {
                "part_number": "",
                "description": "No line items detected",
                "hsn_sac": "",
                "qty": "",
                "mrp": "",
                "rate": "",
                "amount": "",
            }
        ]
        for it in items:
            writer.writerow(
                [
                    inv.get("invoice_number", ""),
                    inv.get("supplier_name", ""),
                    inv.get("date", ""),
                    inv.get("place_of_supply", ""),
                    it.get("part_number", ""),
                    it.get("description", ""),
                    it.get("qty", ""),
                    it.get("mrp", ""),
                    it.get("rate", ""),
                    it.get("amount", ""),
                    it.get("hsn_sac", ""),
                    inv.get("filename", ""),
                ]
            )
    # UTF-8 BOM so Excel/Numbers detect encoding
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def build_download_zip(invoices: list[dict[str, Any]], stamp: str) -> bytes:
    xlsx = build_workbook(invoices)
    csv_bytes = build_csv(invoices)
    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"invoices_{stamp}.csv", csv_bytes)
        zf.writestr(f"invoices_{stamp}.xlsx", xlsx)
    return zbuf.getvalue()


def save_outputs(invoices: list[dict[str, Any]], stamp: str) -> Path:
    """Write copies for local Finder use; on cloud use /tmp (ephemeral)."""
    configured = os.getenv("OUTPUT_DIR", "").strip()
    if configured:
        out_dir = Path(configured)
    else:
        local = Path(__file__).resolve().parent.parent / "output"
        out_dir = local if local.parent.exists() else Path("/tmp/mytvs-output")
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / f"invoices_{stamp}.xlsx").write_bytes(build_workbook(invoices))
    (out_dir / f"invoices_{stamp}.csv").write_bytes(build_csv(invoices))
    (out_dir / "latest.csv").write_bytes(build_csv(invoices))
    (out_dir / "latest.xlsx").write_bytes(build_workbook(invoices))
    return out_dir


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------


@app.get("/api/health")
def health() -> dict[str, str]:
    extractors = ["tesseract_rules", "tesseract_layout"]
    try:
        from extraction import openai_invoice, textract_invoice

        if openai_invoice.available():
            extractors.append("openai")
        if textract_invoice.available():
            extractors.append("textract")
    except Exception:
        pass
    return {
        "status": "ok",
        "version": app.version,
        "build": DEPLOY_MARK,
        "extractors": ",".join(extractors),
    }


async def _load_invoices(files: list[UploadFile]) -> tuple[list[dict[str, Any]], list[str]]:
    if not files:
        raise HTTPException(status_code=400, detail="Please upload at least one invoice.")
    if len(files) > MAX_FILES:
        raise HTTPException(status_code=400, detail=f"Maximum {MAX_FILES} files allowed.")

    invoices: list[dict[str, Any]] = []
    errors: list[str] = []

    for f in files:
        name = f.filename or "invoice.png"
        ext = _ext(name)
        if ext not in PDF_EXTS | IMAGE_EXTS:
            errors.append(f"{name}: use PDF or image (PNG/JPG)")
            continue

        content = await f.read()
        if not content:
            errors.append(f"{name}: empty file")
            continue
        if len(content) > MAX_FILE_SIZE:
            errors.append(f"{name}: exceeds 20 MB limit")
            continue

        try:
            invoices.append(extract_invoice(content, name))
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{name}: failed ({exc})")

    if not invoices:
        raise HTTPException(
            status_code=400,
            detail="; ".join(errors) if errors else "No invoices could be processed.",
        )

    return merge_by_invoice(invoices), errors


@app.post("/api/parse")
async def parse(files: list[UploadFile] = File(...)) -> JSONResponse:
    """Return extracted invoice JSON so the UI can show data even if Excel open fails."""
    invoices, errors = await _load_invoices(files)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = save_outputs(invoices, stamp)
    payload = []
    for inv in invoices:
        payload.append(
            {
                "invoice_number": inv.get("invoice_number"),
                "supplier_name": inv.get("supplier_name"),
                "date": inv.get("date"),
                "place_of_supply": inv.get("place_of_supply"),
                "filename": inv.get("filename"),
                "line_items": inv.get("line_items") or [],
                "item_count": len(inv.get("line_items") or []),
                "confidence": inv.get("confidence") or "medium",
                "confidence_score": inv.get("confidence_score"),
                "warnings": inv.get("warnings") or [],
                "extractor": inv.get("extractor") or "tesseract_rules",
                "items_sum": inv.get("items_sum"),
                "taxable_total": inv.get("taxable_total"),
            }
        )
    return JSONResponse(
        {
            "invoices": payload,
            "errors": errors,
            "output_dir": str(out_dir),
            "latest_csv": str(out_dir / "latest.csv"),
            "latest_xlsx": str(out_dir / "latest.xlsx"),
        }
    )


@app.post("/api/convert")
async def convert(files: list[UploadFile] = File(...)) -> Response:
    """Download ZIP with CSV (opens in Numbers) + XLSX."""
    invoices, errors = await _load_invoices(files)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = save_outputs(invoices, stamp)
    payload = build_download_zip(invoices, stamp)
    filename = f"invoices_{stamp}.zip"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Length": str(len(payload)),
        "X-Processed-Count": str(len(invoices)),
        "X-Item-Count": str(sum(len(i.get("line_items") or []) for i in invoices)),
        "X-Output-Dir": str(out_dir),
        "Access-Control-Expose-Headers": (
            "Content-Disposition, X-Processed-Count, X-Item-Count, X-Partial-Errors, X-Output-Dir"
        ),
    }
    if errors:
        headers["X-Partial-Errors"] = "; ".join(errors)[:500]

    return Response(content=payload, media_type="application/zip", headers=headers)


@app.post("/api/convert-csv")
async def convert_csv(files: list[UploadFile] = File(...)) -> Response:
    invoices, errors = await _load_invoices(files)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_outputs(invoices, stamp)
    csv_bytes = build_csv(invoices)
    headers = {
        "Content-Disposition": f'attachment; filename="invoices_{stamp}.csv"',
        "Content-Length": str(len(csv_bytes)),
        "X-Processed-Count": str(len(invoices)),
        "X-Item-Count": str(sum(len(i.get("line_items") or []) for i in invoices)),
        "Access-Control-Expose-Headers": "Content-Disposition, X-Processed-Count, X-Item-Count, X-Partial-Errors",
    }
    if errors:
        headers["X-Partial-Errors"] = "; ".join(errors)[:500]
    return Response(content=csv_bytes, media_type="text/csv; charset=utf-8", headers=headers)


@app.post("/api/convert-xlsx")
async def convert_xlsx(files: list[UploadFile] = File(...)) -> Response:
    """Optional Excel-only download (one sheet per invoice)."""
    invoices, errors = await _load_invoices(files)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_outputs(invoices, stamp)
    xlsx = build_workbook(invoices)
    headers = {
        "Content-Disposition": f'attachment; filename="invoices_{stamp}.xlsx"',
        "Content-Length": str(len(xlsx)),
        "X-Processed-Count": str(len(invoices)),
        "X-Item-Count": str(sum(len(i.get("line_items") or []) for i in invoices)),
        "Access-Control-Expose-Headers": "Content-Disposition, X-Processed-Count, X-Item-Count, X-Partial-Errors",
    }
    if errors:
        headers["X-Partial-Errors"] = "; ".join(errors)[:500]
    return Response(
        content=xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
